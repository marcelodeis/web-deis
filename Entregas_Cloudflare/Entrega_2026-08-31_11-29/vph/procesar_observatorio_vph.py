#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
=============================================================================
OBSERVATORIO EPIDEMIOLÓGICO Y DASHBOARD PROFESIONAL VPH (2014 - 2026)
SERVICIO DE SALUD OSORNO — PROVINCIA DE OSORNO (7 COMUNAS)
=============================================================================
Motor de Procesamiento de Datos, Analítica Epidemiológica y Exportación.

Reglas Mandatorias MINSAL / DEIS:
  1. VACUNA_ADMINISTRADA == "SI"
  2. REGISTRO_ELIMINADO != "SI" (o == "NO")
  3. Excluir CRITERIO_ELEGIBILIDAD == "EPRO"
  4. Excluir DOSIS == "EPRO"
  5. Criterio exclusivo de RESIDENCIA en el S.S. Osorno (10301-10307)
"""

import csv
import json
import os
import sys
import time
from collections import defaultdict
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: Se requiere openpyxl. Instale con: pip install openpyxl")
    sys.exit(1)

# =============================================================================
# CONFIGURACIÓN GENERAL Y RUTAS
# =============================================================================

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
BD_MINSAL_DIR = os.path.join(os.path.dirname(BASE_DIR), "BASE DATOS MINSAL")

def get_db_dir(year):
    """Devuelve la ruta al directorio de la base de datos según el año"""
    if int(year) <= 2024:
        return os.path.join(BD_MINSAL_DIR, "2000-2024")
    else:
        return os.path.join(BD_MINSAL_DIR, str(year))

OUTPUT_EXCEL = os.path.join(BASE_DIR, "Reporte_Master_Observatorio_VPH_Osorno.xlsx")
OUTPUT_JSON = os.path.join(BASE_DIR, "dashboard_data_vph.json")
OUTPUT_JS = os.path.join(BASE_DIR, "dashboard_data_vph.js")

# Comunas Servicio de Salud Osorno (Código Región 10, Provincia 103)
COMUNAS_OSORNO = {
    "10301": "Osorno",
    "10302": "Puerto Octay",
    "10303": "Purranque",
    "10304": "Puyehue",
    "10305": "Río Negro",
    "10306": "San Juan de la Costa",
    "10307": "San Pablo",
}

CODIGOS_COMUNAS = sorted(COMUNAS_OSORNO.keys())

# Archivos de Programáticas por Residencia (2014 a 2026)
PROGRAMATICAS_FILES = {
    2014: "Programáticas_Residencia_2014",
    2015: "Programáticas_Residencia_2015",
    2016: "Programáticas_Residencia_2016",
    2017: "Programáticas_Residencia_2017",
    2018: "Programáticas_Residencia_2018",
    2019: "Programáticas_Residencia_2019.csv",
    2020: "Programáticas_Residencia_2020.csv",
    2021: "Programáticas_Residencia_2021.csv",
    2022: "Programáticas_Residencia_2022.csv",
    2023: "Programáticas_Residencia_2023.csv",
    2024: "Programáticas_Residencia_2024.csv",
    2025: "Programáticas_Residencia_2025.csv",
    2026: "Programáticas_Residencia_2026.csv",
}

# Años de evaluación del indicador (a los 15 años cumplidos)
# Cohortes 2000 a 2011 -> Evaluación 2015 a 2026
ANOS_EVALUACION = [2015, 2016, 2017, 2018, 2019, 2020, 2021, 2022, 2023, 2024, 2025, 2026]
COHORTES_MAP = {ano: ano - 15 for ano in ANOS_EVALUACION}

VACUNAS_VPH_KEYWORDS = ["VPH", "PAPILOMA", "GARDASIL", "CERVARIX"]


# =============================================================================
# FUNCIONES AUXILIARES DE NORMALIZACIÓN Y FILTRADO
# =============================================================================

def es_vacuna_vph(nombre_vacuna):
    if not nombre_vacuna:
        return False
    nombre_upper = str(nombre_vacuna).upper()
    return any(kw in nombre_upper for kw in VACUNAS_VPH_KEYWORDS)


def clasificar_tipo_vacuna(nombre_vacuna):
    """Clasifica si es Nonavalente (Gardasil 9) o Tetravalente/Bivalente."""
    if not nombre_vacuna:
        return "Desconocida"
    n = str(nombre_vacuna).upper()
    if "NONA" in n or "9" in n or "GARDASIL 9" in n:
        return "VPH Nonavalente (Gardasil 9)"
    if "TETRA" in n or "4" in n or "GARDASIL" in n:
        return "VPH Tetravalente (Gardasil 4)"
    if "CERVARIX" in n or "BIVALENTE" in n or "2" in n:
        return "VPH Bivalente (Cervarix)"
    return "VPH Otra / No Especificada"


def clasificar_tipo_dosis(dosis_str):
    if not dosis_str:
        return "Otra"
    d = str(dosis_str).strip().upper()
    if "ÚNICA" in d or "UNICA" in d:
        return "Dosis Única"
    if "1" in d:
        return "1ª Dosis"
    if "2" in d:
        return "2ª Dosis"
    if "3" in d:
        return "3ª Dosis"
    return "Otra Dosis"


def normalizar_sexo(val):
    if val is None:
        return "Otro"
    v = str(val).strip().upper()
    if v in ("1", "1.0", "H", "HOMBRE", "HOMBRES", "M", "MASCULINO"):
        return "Hombre"
    if v in ("2", "2.0", "M", "MUJER", "MUJERES", "F", "FEMENINO"):
        return "Mujer"
    return "Otro"


def parsear_fecha(fecha_str):
    if not fecha_str:
        return None
    s = str(fecha_str).strip()
    if " " in s:
        s = s.split(" ")[0]
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%d.%m.%Y", "%Y%m%d"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    if len(s) >= 4 and s[:4].isdigit():
        y = int(s[:4])
        if 1990 <= y <= 2030:
            return datetime(y, 1, 1)
    return None


def parsear_ano_nacimiento(fecha_str):
    dt = parsear_fecha(fecha_str)
    if dt:
        return dt.year
    return None


# =============================================================================
# PASO 1: LECTURA DE PROGRAMÁTICAS RNI (2014 - 2026) CON FILTROS MINSAL
# =============================================================================

def leer_programaticas_vph():
    """
    Lee todas las bases de Programáticas por Residencia (2014-2026).
    Aplica rigurosamente los 4 filtros MINSAL y el criterio de residencia.
    """
    print("\n" + "=" * 75)
    print("PASO 1: Lectura de Bases Programáticas RNI (2014-2026) — Criterio Residencia")
    print("=" * 75)

    # Estructura: personas[persona_id] = { datos personales, dosis: [ {...}, ... ] }
    personas = {}
    
    # Producción de dosis anuales: dosis_anuales[ano_adm][comuna][sexo][tipo_dosis] -> count
    dosis_anuales = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(int))))
    
    # Detalle por establecimiento: estab_detalle[estab][tipo_vacuna] -> count
    estab_detalle = defaultdict(lambda: defaultdict(int))
    
    # Matriz completa de ocurrencia por establecimiento: estab_matriz[estab]
    estab_matriz = defaultdict(lambda: {
        "comuna_cod": "10301",
        "comuna_nom": "Osorno",
        "por_ano": defaultdict(lambda: {
            "total": 0,
            "mujeres": 0,
            "hombres": 0,
            "dosis_unica": 0,
            "dosis_1": 0,
            "dosis_2": 0,
            "dosis_otra": 0,
            "tetra": 0,
            "nona": 0,
            "biv": 0,
            "otra_vac": 0
        }),
        "total_historico": 0,
        "total_mujeres": 0,
        "total_hombres": 0,
        "total_tetra": 0,
        "total_nona": 0,
        "total_biv": 0,
        "total_otra_vac": 0,
        "total_dosis_unica": 0,
        "total_dosis_1": 0,
        "total_dosis_2": 0
    })

    # Detalle de vacunas por año de administración: vacunas_por_ano[ano_adm][tipo_vacuna] -> count
    vacunas_por_ano = defaultdict(lambda: defaultdict(int))

    total_registros_vph = 0

    for ano_adm, fname in sorted(PROGRAMATICAS_FILES.items()):
        fpath = os.path.join(get_db_dir(ano_adm), fname)
        if not os.path.exists(fpath):
            print(f"  [AVISO] Archivo no encontrado para año {ano_adm}: {fname}")
            continue

        t0 = time.time()
        count_ano = 0
        print(f"  Leyendo año {ano_adm} ({fname})...")

        with open(fpath, "r", encoding="latin-1", errors="replace") as f:
            first_line = f.readline()
            delim = "|" if "|" in first_line else (";" if ";" in first_line else ",")
            f.seek(0)
            reader = csv.DictReader(f, delimiter=delim)

            for row in reader:
                # Regla 1: VACUNA_ADMINISTRADA == "SI"
                vac_adm = row.get("VACUNA_ADMINISTRADA", "").strip().upper()
                if vac_adm and vac_adm != "SI":
                    continue

                # Regla 2: REGISTRO_ELIMINADO != "SI"
                reg_elim = row.get("REGISTRO_ELIMINADO", "").strip().upper()
                if reg_elim == "SI":
                    continue

                # Regla 3: Excluir CRITERIO_ELEGIBILIDAD == "EPRO"
                crit_eleg = row.get("CRITERIO_ELEGIBILIDAD", "").strip().upper()
                if crit_eleg == "EPRO":
                    continue

                # Regla 4: Excluir DOSIS == "EPRO"
                dosis_raw = row.get("DOSIS", "").strip().upper()
                if dosis_raw == "EPRO":
                    continue

                # Filtro: Solo vacunas VPH
                nom_vac = row.get("VACUNA", "") or row.get("NOMBRE_VACUNA", "") or row.get("VACUNA_DESC", "")
                if not es_vacuna_vph(nom_vac):
                    continue

                # Filtro Criterio Residencia: Comuna de residencia en S.S. Osorno
                com_res = (
                    row.get("COD_COMUNA_RESID", "")
                    or row.get("COD_COMUNA_RESIDENCIA", "")
                    or row.get("COD_COMUNA", "")
                    or row.get("COMUNA_RESIDENCIA", "")
                    or row.get("COMUNA_RESID", "")
                ).strip()

                if not com_res:
                    continue

                if com_res not in COMUNAS_OSORNO:
                    try:
                        com_res = str(int(com_res))
                    except ValueError:
                        pass
                if com_res not in COMUNAS_OSORNO:
                    continue

                # Identificador único de persona
                run = row.get("RUN", "").strip().upper()
                pasaporte = row.get("PASAPORTE", "").strip().upper()
                otro_id = row.get("OTRO", "").strip().upper()
                persona_id = run or pasaporte or otro_id
                if not persona_id:
                    nom = row.get("NOMBRES", "").strip().upper()
                    ap = row.get("APELLIDO_PATERNO", "").strip().upper()
                    fn = row.get("FECHA_NACIMIENTO", "").strip()
                    if nom and ap:
                        persona_id = f"NOM_{nom}_{ap}_{fn}"
                    else:
                        continue

                # Datos del paciente y la dosis
                sexo = normalizar_sexo(row.get("SEXO", ""))
                fnac = row.get("FECHA_NACIMIENTO", "")
                ano_nac = parsear_ano_nacimiento(fnac)
                
                f_vac_str = row.get("FECHA_INMUNIZACION", "") or row.get("FECHA_VACUNACION", "") or row.get("FECHA_ADMINISTRACION", "")
                dt_vac = parsear_fecha(f_vac_str)
                ano_vac = dt_vac.year if dt_vac else ano_adm

                tipo_dosis = clasificar_tipo_dosis(dosis_raw)
                tipo_vac = clasificar_tipo_vacuna(nom_vac)
                estab = (row.get("ESTABLECIMIENTO", "") or row.get("NOMBRE_ESTABLECIMIENTO", "")).strip() or "Establecimiento No Informado"

                # Comuna de ocurrencia / establecimiento
                com_estab = (
                    row.get("COD_COMUNA_ESTAB", "")
                    or row.get("COD_COMUNA_ESTABLECIMIENTO", "")
                    or row.get("COD_COMUNA", "")
                ).strip()
                if not com_estab or com_estab not in COMUNAS_OSORNO:
                    com_estab = com_res

                # Registro de dosis en paciente
                if persona_id not in personas:
                    personas[persona_id] = {
                        "id": persona_id,
                        "sexo": sexo,
                        "ano_nac": ano_nac,
                        "comuna_residencia": com_res,
                        "dosis_list": [],
                    }
                else:
                    if personas[persona_id]["sexo"] == "Otro" and sexo in ("Hombre", "Mujer"):
                        personas[persona_id]["sexo"] = sexo
                    if not personas[persona_id]["ano_nac"] and ano_nac:
                        personas[persona_id]["ano_nac"] = ano_nac
                    personas[persona_id]["comuna_residencia"] = com_res

                personas[persona_id]["dosis_list"].append({
                    "ano_adm": ano_adm,
                    "ano_vac": ano_vac,
                    "tipo_dosis": tipo_dosis,
                    "tipo_vacuna": tipo_vac,
                    "nombre_vacuna": nom_vac,
                    "comuna": com_res,
                    "establecimiento": estab,
                })

                # Estadísticas de producción de dosis
                dosis_anuales[ano_adm][com_res][sexo][tipo_dosis] += 1
                dosis_anuales[ano_adm][com_res][sexo]["Total"] += 1
                dosis_anuales[ano_adm][com_res]["Total"][tipo_dosis] += 1
                dosis_anuales[ano_adm][com_res]["Total"]["Total"] += 1
                dosis_anuales[ano_adm]["TOTAL"][sexo][tipo_dosis] += 1
                dosis_anuales[ano_adm]["TOTAL"][sexo]["Total"] += 1
                dosis_anuales[ano_adm]["TOTAL"]["Total"][tipo_dosis] += 1
                dosis_anuales[ano_adm]["TOTAL"]["Total"]["Total"] += 1

                estab_detalle[estab][tipo_vac] += 1
                vacunas_por_ano[ano_adm][tipo_vac] += 1

                # Actualizar matriz completa de ocurrencia por establecimiento
                em = estab_matriz[estab]
                em["comuna_cod"] = com_estab
                em["comuna_nom"] = COMUNAS_OSORNO.get(com_estab, "Osorno")
                
                em_ano = em["por_ano"][ano_adm]
                em_ano["total"] += 1
                if sexo == "Mujer":
                    em_ano["mujeres"] += 1
                    em["total_mujeres"] += 1
                elif sexo == "Hombre":
                    em_ano["hombres"] += 1
                    em["total_hombres"] += 1

                if tipo_dosis == "Dosis Única":
                    em_ano["dosis_unica"] += 1
                    em["total_dosis_unica"] += 1
                elif tipo_dosis == "1ª Dosis":
                    em_ano["dosis_1"] += 1
                    em["total_dosis_1"] += 1
                elif tipo_dosis == "2ª Dosis":
                    em_ano["dosis_2"] += 1
                    em["total_dosis_2"] += 1
                else:
                    em_ano["dosis_otra"] += 1

                if tipo_vac == "VPH Tetravalente (Gardasil 4)":
                    em_ano["tetra"] += 1
                    em["total_tetra"] += 1
                elif tipo_vac == "VPH Nonavalente (Gardasil 9)":
                    em_ano["nona"] += 1
                    em["total_nona"] += 1
                elif tipo_vac == "VPH Bivalente (Cervarix)":
                    em_ano["biv"] += 1
                    em["total_biv"] += 1
                else:
                    em_ano["otra_vac"] += 1
                    em["total_otra_vac"] += 1

                em["total_historico"] += 1

                count_ano += 1
                total_registros_vph += 1

        elapsed = time.time() - t0
        print(f"    -> {count_ano:,} registros VPH válidos procesados ({elapsed:.1f}s)")

    print(f"\n  [OK] Total de personas únicas con registro VPH: {len(personas):,}")
    print(f"  [OK] Total de registros de dosis VPH válidas: {total_registros_vph:,}")

    return personas, dosis_anuales, estab_detalle, vacunas_por_ano, estab_matriz


# =============================================================================
# PASO 2: EVALUACIÓN DE ESQUEMA COMPLETO Y PERFIL EPIDEMIOLÓGICO
# =============================================================================

def determinar_esquema_completo(dosis_list):
    """
    Determina si la persona tiene esquema completo según la normativa PNI:
      - 1+ dosis de VPH Nonavalente (Gardasil 9) -> Completo
      - 1+ dosis clasificada como 'Dosis Única' -> Completo
      - 2+ dosis registradas de cualquier vacuna VPH -> Completo
      - De lo contrario -> Incompleto (1 dosis tetravalente/bivalente)
    """
    if not dosis_list:
        return False, "Sin Dosis", 0

    n_dosis = len(dosis_list)
    tiene_nonavalente = any("NONA" in d.get("tipo_vacuna", "").upper() or "9" in d.get("tipo_vacuna", "") for d in dosis_list)
    tiene_dosis_unica = any(d.get("tipo_dosis") == "Dosis Única" for d in dosis_list)

    if tiene_nonavalente:
        return True, "Dosis Única (Nonavalente)", n_dosis
    if tiene_dosis_unica:
        return True, "Dosis Única", n_dosis
    if n_dosis >= 2:
        return True, "2+ Dosis (Esquema Multidosis)", n_dosis

    return False, "1 Dosis Incompleta", n_dosis


def procesar_cohortes_vacunadas(personas):
    """
    Agrupa los vacunados con esquema completo e incompleto por cohorte y comuna.
    Retorna: vacunados_cohorte[cohorte][comuna][sexo] = { completo, incompleto, total_iniciados }
    """
    print("\n" + "=" * 75)
    print("PASO 2: Evaluación de Esquema Completo por Persona, Cohorte y Sexo")
    print("=" * 75)

    vacunados_cohorte = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: {
        "completo": 0,
        "incompleto": 0,
        "iniciados": 0,
        "tipo_esquema": defaultdict(int),
    })))

    for pid, pdata in personas.items():
        ano_nac = pdata.get("ano_nac")
        if not ano_nac or ano_nac < 1995 or ano_nac > 2025:
            continue

        comuna = pdata.get("comuna_residencia", "")
        if comuna not in COMUNAS_OSORNO:
            continue

        sexo = pdata.get("sexo", "Mujer")
        if sexo not in ("Hombre", "Mujer"):
            sexo = "Mujer"

        es_completo, tipo_esq, n_dosis = determinar_esquema_completo(pdata["dosis_list"])

        # Actualizar conteos por comuna
        entry_com = vacunados_cohorte[ano_nac][comuna][sexo]
        entry_com["iniciados"] += 1
        entry_com["tipo_esquema"][tipo_esq] += 1
        if es_completo:
            entry_com["completo"] += 1
        else:
            entry_com["incompleto"] += 1

        # Totales comunales (ambos sexos)
        entry_com_tot = vacunados_cohorte[ano_nac][comuna]["Total"]
        entry_com_tot["iniciados"] += 1
        entry_com_tot["tipo_esquema"][tipo_esq] += 1
        if es_completo:
            entry_com_tot["completo"] += 1
        else:
            entry_com_tot["incompleto"] += 1

        # Totales provinciales por sexo
        entry_prov_sex = vacunados_cohorte[ano_nac]["TOTAL"][sexo]
        entry_prov_sex["iniciados"] += 1
        entry_prov_sex["tipo_esquema"][tipo_esq] += 1
        if es_completo:
            entry_prov_sex["completo"] += 1
        else:
            entry_prov_sex["incompleto"] += 1

        # Total provincial general
        entry_prov_tot = vacunados_cohorte[ano_nac]["TOTAL"]["Total"]
        entry_prov_tot["iniciados"] += 1
        entry_prov_tot["tipo_esquema"][tipo_esq] += 1
        if es_completo:
            entry_prov_tot["completo"] += 1
        else:
            entry_prov_tot["incompleto"] += 1

    for ano_med in ANOS_EVALUACION:
        coh = COHORTES_MAP[ano_med]
        tot = vacunados_cohorte[coh]["TOTAL"]["Total"]["completo"]
        fem = vacunados_cohorte[coh]["TOTAL"]["Mujer"]["completo"]
        masc = vacunados_cohorte[coh]["TOTAL"]["Hombre"]["completo"]
        print(f"  Año {ano_med} (Cohorte {coh}): Esquema Completo Total={tot:,} (Mujeres={fem:,}, Hombres={masc:,})")

    return vacunados_cohorte


# =============================================================================
# PASO 3: LECTURA DE NACIDOS VIVOS DEIS (2000 - 2021)
# =============================================================================

def leer_nacimientos_por_cohorte():
    """
    Lee archivos NAC2000 a NAC2021 de DEIS por residencia.
    Retorna: nacimientos[ano_nac][comuna][sexo] -> count
    """
    print("\n" + "=" * 75)
    print("PASO 3: Lectura de Nacidos Vivos DEIS (Criterio Residencia y Sexo)")
    print("=" * 75)

    nacimientos = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
    anos_necesarios = sorted(set(COHORTES_MAP.values()))

    for ano_nac in anos_necesarios:
        f_xlsx = os.path.join(get_db_dir(ano_nac), f"NAC{ano_nac}.xlsx")
        f_csv = os.path.join(get_db_dir(ano_nac), f"NAC{ano_nac}.csv")

        if os.path.exists(f_xlsx):
            print(f"  Leyendo nacimientos {ano_nac} (Excel)...")
            wb = openpyxl.load_workbook(f_xlsx, read_only=True, data_only=True)
            ws = wb.active
            header = [str(c.value).strip().upper() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

            idx_com = header.index("COMUNA") if "COMUNA" in header else (header.index("COD_COMUNA") if "COD_COMUNA" in header else None)
            idx_sexo = header.index("SEXO") if "SEXO" in header else next((i for i, h in enumerate(header) if "SEX" in h), None)

            if idx_com is not None:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if len(row) > idx_com and row[idx_com] is not None:
                        try:
                            val = int(row[idx_com]) if isinstance(row[idx_com], (int, float)) else int(str(row[idx_com]).strip())
                            com_str = str(val)
                        except (ValueError, TypeError):
                            continue
                        if com_str in COMUNAS_OSORNO:
                            sexo = "Otro"
                            if idx_sexo is not None and len(row) > idx_sexo and row[idx_sexo] is not None:
                                sexo = normalizar_sexo(row[idx_sexo])
                            if sexo not in ("Hombre", "Mujer"):
                                sexo = "Mujer"

                            nacimientos[ano_nac][com_str][sexo] += 1
                            nacimientos[ano_nac][com_str]["Total"] += 1
                            nacimientos[ano_nac]["TOTAL"][sexo] += 1
                            nacimientos[ano_nac]["TOTAL"]["Total"] += 1
            wb.close()

        elif os.path.exists(f_csv):
            print(f"  Leyendo nacimientos {ano_nac} (CSV)...")
            with open(f_csv, "r", encoding="latin-1", errors="replace") as f:
                first = f.readline()
                delim = ";" if ";" in first else ("," if "," in first else "|")
                f.seek(0)
                reader = csv.reader(f, delimiter=delim)
                header = [h.strip().upper() for h in next(reader)]

                idx_com = header.index("COMUNA") if "COMUNA" in header else (header.index("COD_COMUNA") if "COD_COMUNA" in header else None)
                idx_sexo = header.index("SEXO") if "SEXO" in header else next((i for i, h in enumerate(header) if "SEX" in h), None)

                if idx_com is not None:
                    for row in reader:
                        if len(row) > idx_com:
                            val = row[idx_com].strip()
                            if val.isdigit():
                                com_str = str(int(val))
                                if com_str in COMUNAS_OSORNO:
                                    sexo = "Otro"
                                    if idx_sexo is not None and len(row) > idx_sexo:
                                        sexo = normalizar_sexo(row[idx_sexo].strip())
                                    if sexo not in ("Hombre", "Mujer"):
                                        sexo = "Mujer"

                                    nacimientos[ano_nac][com_str][sexo] += 1
                                    nacimientos[ano_nac][com_str]["Total"] += 1
                                    nacimientos[ano_nac]["TOTAL"][sexo] += 1
                                    nacimientos[ano_nac]["TOTAL"]["Total"] += 1

        tot = nacimientos[ano_nac]["TOTAL"]["Total"]
        fem = nacimientos[ano_nac]["TOTAL"]["Mujer"]
        masc = nacimientos[ano_nac]["TOTAL"]["Hombre"]
        print(f"    -> Cohorte {ano_nac}: Total={tot:,} (Mujeres={fem:,}, Hombres={masc:,})")

    return nacimientos


# =============================================================================
# PASO 4: LECTURA DE DEFUNCIONES DEIS (1999 - 2026)
# =============================================================================

def leer_defunciones_por_cohorte():
    """
    Lee archivos DEF1999 a DEF2026 de DEIS por residencia.
    Filtra defunciones ocurridas a menores de 15 años.
    Retorna: defunciones[cohorte_nac][comuna][sexo] -> count
    """
    print("\n" + "=" * 75)
    print("PASO 4: Lectura de Defunciones DEIS 0-14 años (Criterio Residencia y Sexo)")
    print("=" * 75)

    defunciones = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
    anos_cohortes = set(COHORTES_MAP.values())

    for ano_def in range(1999, 2027):
        f_xlsx = os.path.join(get_db_dir(ano_def), f"DEF{ano_def}.xlsx")
        f_csv = os.path.join(get_db_dir(ano_def), f"DEF{ano_def}.csv")

        if os.path.exists(f_xlsx):
            print(f"  Leyendo defunciones {ano_def} (Excel)...")
            wb = openpyxl.load_workbook(f_xlsx, read_only=True, data_only=True)
            ws = wb.active
            header = [str(c.value).strip().upper() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

            if "ANO1_NAC" in header and "ANO2_NAC" in header and "EDAD_TIPO" in header and "EDAD_CANT" in header:
                idx_a1 = header.index("ANO1_NAC")
                idx_a2 = header.index("ANO2_NAC")
                idx_t = header.index("EDAD_TIPO")
                idx_c = header.index("EDAD_CANT")
                idx_s = header.index("SEXO") if "SEXO" in header else next((i for i, h in enumerate(header) if "SEX" in h), None)
                idx_com = header.index("COMUNA") if "COMUNA" in header else (header.index("RES_COMUNA") if "RES_COMUNA" in header else (header.index("COD_COMUNA") if "COD_COMUNA" in header else None))

                for row in ws.iter_rows(min_row=2, values_only=True):
                    if len(row) > max(idx_a1, idx_a2, idx_t, idx_c):
                        try:
                            a1 = int(row[idx_a1]) if row[idx_a1] is not None else 0
                            a2 = int(row[idx_a2]) if row[idx_a2] is not None else 0
                            anac = a1 * 100 + a2
                        except (ValueError, TypeError):
                            continue

                        if anac not in anos_cohortes:
                            continue

                        try:
                            t = int(row[idx_t]) if row[idx_t] is not None else 0
                            c = int(row[idx_c]) if row[idx_c] is not None else 0
                        except (ValueError, TypeError):
                            continue

                        es_menor_15 = (t == 1 and c < 15) or (t in (2, 3, 4, 5))
                        if not es_menor_15:
                            continue

                        com_str = ""
                        if idx_com is not None and len(row) > idx_com and row[idx_com] is not None:
                            try:
                                com_str = str(int(row[idx_com])) if isinstance(row[idx_com], (int, float)) else str(row[idx_com]).strip()
                            except (ValueError, TypeError):
                                pass

                        if com_str in COMUNAS_OSORNO:
                            sexo = "Mujer"
                            if idx_s is not None and len(row) > idx_s and row[idx_s] is not None:
                                sexo = normalizar_sexo(row[idx_s])
                            if sexo not in ("Hombre", "Mujer"):
                                sexo = "Mujer"

                            defunciones[anac][com_str][sexo] += 1
                            defunciones[anac][com_str]["Total"] += 1
                            defunciones[anac]["TOTAL"][sexo] += 1
                            defunciones[anac]["TOTAL"]["Total"] += 1
            wb.close()

        elif os.path.exists(f_csv):
            print(f"  Leyendo defunciones {ano_def} (CSV)...")
            with open(f_csv, "r", encoding="latin-1", errors="replace") as f:
                first = f.readline()
                delim = ";" if ";" in first else ("," if "," in first else "|")
                f.seek(0)
                reader = csv.reader(f, delimiter=delim)
                header = [h.strip().upper() for h in next(reader)]

                if "ANO1_NAC" in header and "ANO2_NAC" in header and "EDAD_TIPO" in header and "EDAD_CANT" in header:
                    idx_a1 = header.index("ANO1_NAC")
                    idx_a2 = header.index("ANO2_NAC")
                    idx_t = header.index("EDAD_TIPO")
                    idx_c = header.index("EDAD_CANT")
                    idx_s = header.index("SEXO") if "SEXO" in header else next((i for i, h in enumerate(header) if "SEX" in h), None)
                    idx_com = header.index("COMUNA") if "COMUNA" in header else (header.index("RES_COMUNA") if "RES_COMUNA" in header else (header.index("COD_COMUNA") if "COD_COMUNA" in header else None))

                    for row in reader:
                        if len(row) > max(idx_a1, idx_a2, idx_t, idx_c):
                            try:
                                a1 = int(row[idx_a1].strip())
                                a2 = int(row[idx_a2].strip())
                                anac = a1 * 100 + a2
                            except (ValueError, TypeError):
                                continue

                            if anac not in anos_cohortes:
                                continue

                            try:
                                t = int(row[idx_t].strip())
                                c = int(row[idx_c].strip())
                            except (ValueError, TypeError):
                                continue

                            es_menor_15 = (t == 1 and c < 15) or (t in (2, 3, 4, 5))
                            if not es_menor_15:
                                continue

                            com_str = ""
                            if idx_com is not None and len(row) > idx_com:
                                try:
                                    com_str = str(int(row[idx_com].strip()))
                                except (ValueError, TypeError):
                                    pass

                            if com_str in COMUNAS_OSORNO:
                                sexo = "Mujer"
                                if idx_s is not None and len(row) > idx_s:
                                    sexo = normalizar_sexo(row[idx_s].strip())
                                if sexo not in ("Hombre", "Mujer"):
                                    sexo = "Mujer"

                                defunciones[anac][com_str][sexo] += 1
                                defunciones[anac][com_str]["Total"] += 1
                                defunciones[anac]["TOTAL"][sexo] += 1
                                defunciones[anac]["TOTAL"]["Total"] += 1

    for coh in sorted(anos_cohortes):
        tot = defunciones[coh]["TOTAL"]["Total"]
        fem = defunciones[coh]["TOTAL"]["Mujer"]
        masc = defunciones[coh]["TOTAL"]["Hombre"]
        print(f"  Defunciones 0-14 años cohorte {coh}: Total={tot} (Mujeres={fem}, Hombres={masc})")

    return defunciones


# =============================================================================
# PASO 5: CONSOLIDACIÓN Y CÁLCULO DE MÉTRICAS EPIDEMIOLÓGICAS
# =============================================================================

def calcular_indicadores_completos(vacunados, nacimientos, defunciones, dosis_anuales):
    """
    Consolida numeradores, denominadores, coberturas, drop-out rates,
    razón de paridad y coeficientes de disparidad territorial.
    """
    print("\n" + "=" * 75)
    print("PASO 5: Consolidación y Cálculo de Métricas Epidemiológicas")
    print("=" * 75)

    indicadores = {}

    for ano_med in ANOS_EVALUACION:
        coh = COHORTES_MAP[ano_med]
        indicadores[ano_med] = {
            "ano_medicion": ano_med,
            "cohorte": coh,
            "comunas": {},
            "provincial": {},
            "metricas_territoriales": {},
        }

        # Cálculo por Comuna
        cob_list = []
        for com_cod in CODIGOS_COMUNAS:
            com_nom = COMUNAS_OSORNO[com_cod]
            indicadores[ano_med]["comunas"][com_cod] = {
                "nombre": com_nom,
                "Total": {"num": 0, "iniciados": 0, "den": 0, "cob": 0.0, "nac": 0, "def": 0, "dropout": 0.0},
                "Mujer": {"num": 0, "iniciados": 0, "den": 0, "cob": 0.0, "nac": 0, "def": 0, "dropout": 0.0},
                "Hombre": {"num": 0, "iniciados": 0, "den": 0, "cob": 0.0, "nac": 0, "def": 0, "dropout": 0.0},
            }

            for sexo in ("Total", "Mujer", "Hombre"):
                num = vacunados[coh][com_cod][sexo]["completo"]
                iniciados = vacunados[coh][com_cod][sexo]["iniciados"]
                nac = nacimientos[coh][com_cod][sexo]
                deff = defunciones[coh][com_cod][sexo]
                den = max(0, nac - deff)
                cob = (num / den * 100.0) if den > 0 else 0.0

                # Tasa de abandono (drop-out de 1ª a 2ª dosis)
                dropout = ((iniciados - num) / iniciados * 100.0) if iniciados > 0 else 0.0

                indicadores[ano_med]["comunas"][com_cod][sexo] = {
                    "num": num,
                    "iniciados": iniciados,
                    "den": den,
                    "cob": round(cob, 2),
                    "nac": nac,
                    "def": deff,
                    "dropout": round(max(0.0, dropout), 2),
                }

            cob_list.append(indicadores[ano_med]["comunas"][com_cod]["Total"]["cob"])

        # Cálculo Provincial
        for sexo in ("Total", "Mujer", "Hombre"):
            num = vacunados[coh]["TOTAL"][sexo]["completo"]
            iniciados = vacunados[coh]["TOTAL"][sexo]["iniciados"]
            nac = nacimientos[coh]["TOTAL"][sexo]
            deff = defunciones[coh]["TOTAL"][sexo]
            den = max(0, nac - deff)
            cob = (num / den * 100.0) if den > 0 else 0.0
            dropout = ((iniciados - num) / iniciados * 100.0) if iniciados > 0 else 0.0

            indicadores[ano_med]["provincial"][sexo] = {
                "num": num,
                "iniciados": iniciados,
                "den": den,
                "cob": round(cob, 2),
                "nac": nac,
                "def": deff,
                "dropout": round(max(0.0, dropout), 2),
            }

        # Razón de Paridad de Género (GPI = Cobertura Mujeres / Cobertura Hombres)
        cob_m = indicadores[ano_med]["provincial"]["Mujer"]["cob"]
        cob_h = indicadores[ano_med]["provincial"]["Hombre"]["cob"]
        gpi = (cob_m / cob_h) if cob_h > 0 else None

        # Dispersión Territorial
        max_cob = max(cob_list) if cob_list else 0.0
        min_cob = min(cob_list) if cob_list else 0.0
        mean_cob = sum(cob_list) / len(cob_list) if cob_list else 0.0
        variance = sum((x - mean_cob) ** 2 for x in cob_list) / len(cob_list) if cob_list else 0.0
        std_dev = variance ** 0.5
        cv = (std_dev / mean_cob * 100.0) if mean_cob > 0 else 0.0

        indicadores[ano_med]["metricas_territoriales"] = {
            "gpi": round(gpi, 2) if gpi else None,
            "max_cob": round(max_cob, 2),
            "min_cob": round(min_cob, 2),
            "brecha_max_min": round(max_cob - min_cob, 2),
            "media_comunal": round(mean_cob, 2),
            "desv_est": round(std_dev, 2),
            "cv_territorial": round(cv, 2),
        }

        print(f"\nAÑO {ano_med} (Cohorte {coh}):")
        print(f"  TOTAL S.S. OSORNO: Num={indicadores[ano_med]['provincial']['Total']['num']:,}, Den={indicadores[ano_med]['provincial']['Total']['den']:,} -> {indicadores[ano_med]['provincial']['Total']['cob']}%")
        print(f"  - Mujeres:         Num={indicadores[ano_med]['provincial']['Mujer']['num']:,}, Den={indicadores[ano_med]['provincial']['Mujer']['den']:,} -> {indicadores[ano_med]['provincial']['Mujer']['cob']}%")
        print(f"  - Hombres:         Num={indicadores[ano_med]['provincial']['Hombre']['num']:,}, Den={indicadores[ano_med]['provincial']['Hombre']['den']:,} -> {indicadores[ano_med]['provincial']['Hombre']['cob']}%")
        print(f"  - Brecha Territorial: Max={max_cob}%, Min={min_cob}%, CV={round(cv, 1)}%")

    return indicadores


# =============================================================================
# PASO 6: EXPORTACIÓN PARA DASHBOARD WEB (JSON Y JS STANDALONE)
# =============================================================================

def get_file_mtime_formatted(file_path):
    try:
        mtime = os.path.getmtime(file_path)
        return datetime.fromtimestamp(mtime).strftime("%d/%m/%Y %H:%M")
    except:
        return datetime.now().strftime("%d/%m/%Y %H:%M")

def exportar_datos_dashboard(indicadores, dosis_anuales, estab_detalle, vacunas_por_ano, estab_matriz=None):
    """
    Genera los archivos dashboard_data_vph.json y dashboard_data_vph.js
    para alimentar la interfaz web interactiva.
    """
    print("\n" + "=" * 75)
    print("PASO 6: Exportación de Datos para Dashboard Web")
    print("=" * 75)

    # Formatear matriz completa de establecimientos para filtrado y exportación en web
    matriz_estab_list = []
    if estab_matriz:
        for est, d in sorted(estab_matriz.items(), key=lambda x: x[1]["total_historico"], reverse=True):
            matriz_estab_list.append({
                "establecimiento": est,
                "comuna_cod": d["comuna_cod"],
                "comuna_nom": d["comuna_nom"],
                "total_historico": d["total_historico"],
                "total_mujeres": d["total_mujeres"],
                "total_hombres": d["total_hombres"],
                "total_tetra": d["total_tetra"],
                "total_nona": d["total_nona"],
                "total_biv": d["total_biv"],
                "total_otra_vac": d["total_otra_vac"],
                "total_dosis_unica": d["total_dosis_unica"],
                "total_dosis_1": d["total_dosis_1"],
                "total_dosis_2": d["total_dosis_2"],
                "por_ano": {
                    str(ano): dict(ano_d) for ano, ano_d in d["por_ano"].items()
                }
            })

    data_payload = {
        "metadata": {
            "titulo": "Observatorio Epidemiológico de Vacunación VPH",
            "subtitulo": "Servicio de Salud Osorno — Serie Histórica y Monitoreo Actual (2014-2026)",
            "fecha_corte": get_file_mtime_formatted(os.path.join(get_db_dir(2026), PROGRAMATICAS_FILES[2026])),
            "nota_corte_2026": "Corte preliminar de base RNI 2026 (datos en desarrollo continuo durante el año escolar).",
            "fuentes": "Registro Nacional de Inmunizaciones (RNI) y Departamento de Estadísticas e Información de Salud (DEIS) — MINSAL Chile.",
            "criterio": "Criterio de Residencia Comunal en la Provincia de Osorno (7 comunas).",
            "filtros_aplicados": [
                "VACUNA_ADMINISTRADA == 'SI'",
                "REGISTRO_ELIMINADO != 'SI'",
                "Exclusión de CRITERIO_ELEGIBILIDAD == 'EPRO'",
                "Exclusión de DOSIS == 'EPRO'",
                "Residencia en comunas: 10301 a 10307",
            ],
            "comunas": COMUNAS_OSORNO,
        },
        "hitos_normativos": [
            {
                "ano": 2014,
                "titulo": "Inicio de Vacunación VPH en Chile",
                "descripcion": "Introducción de la vacuna VPH Tetravalente (Gardasil 4) en niñas de 4° Básico (1ª Dosis escolar).",
                "poblacion": "Solo Mujeres (4° Básico)",
                "esquema": "2 Dosis (0-12 meses)",
                "vacuna": "Tetravalente (Gardasil 4)",
                "fuente_bibliografica": "Decreto Exento N° 1201 (2013/2014) MINSAL · Norma Técnica Nacional del Programa Nacional de Inmunizaciones (PNI).",
                "cita_tipo": "Decreto Supremo / Norma Oficial"
            },
            {
                "ano": 2015,
                "titulo": "Consolidación Esquema 2 Dosis en Mujeres",
                "descripcion": "Vacunación a 4° Básico (1ª dosis) y 5° Básico (2ª dosis a las vacunadas en 2014).",
                "poblacion": "Solo Mujeres (4° y 5° Básico)",
                "esquema": "2 Dosis",
                "vacuna": "Tetravalente (Gardasil 4)",
                "fuente_bibliografica": "Ordinario B27 N° 2374 / MINSAL (2015) · Lineamientos Técnicos y Operativos de Vacunación en Establecimientos Educacionales.",
                "cita_tipo": "Lineamiento Técnico MINSAL"
            },
            {
                "ano": 2019,
                "titulo": "Universalización: Incorporación de Varones",
                "descripcion": "Se incorpora a los niños de 4° básico al programa nacional de vacunación VPH escolar.",
                "poblacion": "Hombres y Mujeres (4° y 5° Básico)",
                "esquema": "2 Dosis",
                "vacuna": "Tetravalente (Gardasil 4)",
                "fuente_bibliografica": "Decreto Exento N° 6 (2019) MINSAL · Recomendación Oficial del Comité Asesor en Vacunas e Inmunizaciones (CAVEI 2018).",
                "cita_tipo": "Decreto Supremo MINSAL"
            },
            {
                "ano": 2020,
                "titulo": "Pandemia COVID-19 y Estrategias de Rescate",
                "descripcion": "Cierre de escuelas impacta la vacunación escolar; se implementan operativos de rescate en centros de salud y cursos superiores (6°, 7° y 8° básico).",
                "poblacion": "Ambos sexos escolar y catch-up",
                "esquema": "2 Dosis",
                "vacuna": "Tetravalente (Gardasil 4)",
                "fuente_bibliografica": "Circular C37 N° 08 / MINSAL (2020) · Plan Extraordinario de Puesta al Día de Inmunización Escolar en APS.",
                "cita_tipo": "Circular Ministerial"
            },
            {
                "ano": 2023,
                "titulo": "Transición hacia Esquema de Dosis Única",
                "descripcion": "Chile adopta las recomendaciones de la OMS y NITAG iniciando la transición a Dosis Única con VPH Nonavalente.",
                "poblacion": "Ambos sexos",
                "esquema": "Transición a Dosis Única",
                "vacuna": "Transición a Nonavalente (Gardasil 9)",
                "fuente_bibliografica": "Documento de Posición OMS/SAGE: WER 2022, 97:645-672 · Ensayos KEN SHE (NEJM 2022) & DoRIS (Lancet Oncol 2022) · Dictamen CAVEI 2023.",
                "cita_tipo": "Evidencia Científica OMS / CAVEI"
            },
            {
                "ano": 2024,
                "titulo": "Implementación Plena de Dosis Única Nonavalente",
                "descripcion": "Se establece formalmente el esquema de Dosis Única universal con vacuna VPH Nonavalente (Gardasil 9) en 4° Básico.",
                "poblacion": "Ambos sexos (4° Básico)",
                "esquema": "Dosis Única (1 dosis = completo)",
                "vacuna": "Nonavalente (Gardasil 9)",
                "fuente_bibliografica": "Decreto Exento N° 11 (2024) MINSAL · Norma General Técnica N° 0229 del PNI y Lineamientos Operativos Campaña Escolar 2024.",
                "cita_tipo": "Decreto Supremo MINSAL"
            },
            {
                "ano": 2026,
                "titulo": "Monitoreo en Desarrollo (Campaña 2026)",
                "descripcion": "Campaña escolar en curso con esquema de Dosis Única Nonavalente para ambos sexos.",
                "poblacion": "Ambos sexos (4° Básico y rescate)",
                "esquema": "Dosis Única",
                "vacuna": "Nonavalente (Gardasil 9)",
                "fuente_bibliografica": "Calendario Nacional de Vacunación 2026 · Subsecretaría de Salud Pública, DIPRECE – MINSAL.",
                "cita_tipo": "Calendario Oficial MINSAL"
            },
        ],
        "indicadores_anuales": indicadores,
        "dosis_anuales": dosis_anuales,
        "top_establecimientos": [
            {"nombre": est, "vacunas": dict(vacs), "total": sum(vacs.values())}
            for est, vacs in sorted(estab_detalle.items(), key=lambda x: sum(x[1].values()), reverse=True)[:30]
        ],
        "matriz_establecimientos": matriz_estab_list,
        "evolucion_vacunas": {ano: dict(vacs) for ano, vacs in sorted(vacunas_por_ano.items())},
    }

    # Guardar JSON
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(data_payload, f, ensure_ascii=False, indent=2)
    print(f"  [OK] Archivo JSON exportado: {OUTPUT_JSON}")

    # Guardar JS (para uso local sin problemas de CORS)
    js_content = f"// Datos autogenerados del Observatorio VPH S.S. Osorno\nwindow.VPH_DASHBOARD_DATA = {json.dumps(data_payload, ensure_ascii=False, indent=2)};\n"
    with open(OUTPUT_JS, "w", encoding="utf-8") as f:
        f.write(js_content)
    print(f"  [OK] Archivo JS exportado: {OUTPUT_JS}")


# =============================================================================
# PASO 7: REPORTE MAESTRO EXCEL PROFESIONAL
# =============================================================================

def crear_reporte_excel_maestro(indicadores, dosis_anuales, estab_detalle, vacunas_por_ano):
    """
    Genera un archivo Excel Maestro con 7 hojas formateadas profesionalmente:
      1. Resumen Ejecutivo y Ficha Técnica
      2. Indicador 15 Años — Total Provincial y Comunal
      3. Indicador 15 Años — Mujeres
      4. Indicador 15 Años — Hombres
      5. Dosis Anuales por Comuna y Sexo (2014-2026)
      6. Denominadores DEIS (Nacimientos y Defunciones)
      7. Detalle por Establecimiento de Salud
    """
    print("\n" + "=" * 75)
    print("PASO 7: Generación de Reporte Maestro Excel")
    print("=" * 75)

    wb = openpyxl.Workbook()
    # Eliminar hoja por defecto si existe
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Estilos institucionales
    font_title = Font(name="Calibri", size=15, bold=True, color="FFFFFF")
    font_subtitle = Font(name="Calibri", size=10, italic=True, color="D0E1FD")
    font_sec = Font(name="Calibri", size=11, bold=True, color="1E3A8A")
    font_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=10, bold=True)
    font_regular = Font(name="Calibri", size=10)

    fill_primary = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")     # Azul Marino MINSAL
    fill_secondary = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")   # Azul Medio
    fill_accent = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")      # Teal / Verde Agua
    fill_fem = PatternFill(start_color="9333EA", end_color="9333EA", fill_type="solid")         # Púrpura Mujeres
    fill_masc = PatternFill(start_color="0284C7", end_color="0284C7", fill_type="solid")        # Celeste Hombres
    fill_total_row = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")   # Azul muy suave
    fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")       # Gris tenue

    border_thin = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    border_double_bottom = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="double", color="1E3A8A"),
    )

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    def auto_fit_columns(ws, max_cols=30):
        for col_idx in range(1, min(ws.max_column + 1, max_cols)):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row in range(1, min(ws.max_row + 1, 100)):
                val = ws.cell(row=row, column=col_idx).value
                if val:
                    lines = str(val).split("\n")
                    for line in lines:
                        if len(line) < 60:
                            max_len = max(max_len, len(line))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # -------------------------------------------------------------------------
    # HOJA 1: RESUMEN Y FICHA TÉCNICA
    # -------------------------------------------------------------------------
    ws1 = wb.create_sheet(title="Resumen y Ficha Técnica")
    ws1.views.sheetView[0].showGridLines = True

    ws1.merge_cells("A1:I2")
    ws1["A1"] = "OBSERVATORIO EPIDEMIOLÓGICO DE VACUNACIÓN VPH (2014 - 2026)"
    ws1["A1"].font = font_title
    ws1["A1"].fill = fill_primary
    ws1["A1"].alignment = align_center

    ws1.merge_cells("A3:I3")
    ws1["A3"] = "SERVICIO DE SALUD OSORNO — MONITOREO HISTÓRICO Y EVALUACIÓN A LOS 15 AÑOS (CRITERIO RESIDENCIA)"
    ws1["A3"].font = font_subtitle
    ws1["A3"].fill = fill_secondary
    ws1["A3"].alignment = align_center

    # Tabla Resumen de Coberturas a los 15 Años (2015-2026)
    ws1.cell(row=5, column=1, value="1. RESUMEN PROVINCIAL DE COBERTURA A LOS 15 AÑOS (S.S. OSORNO)").font = font_sec
    
    headers_resumen = [
        "Año Medición", "Cohorte Nac.", "Numerador (Esquema Completo)", "Población Objetivo (DEIS)",
        "Cobertura Total (%)", "Cobertura Mujeres (%)", "Cobertura Hombres (%)", "Razón Paridad (M/H)", "Brecha Territorial (%)"
    ]
    for col_idx, h in enumerate(headers_resumen, start=1):
        cell = ws1.cell(row=6, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_primary
        cell.alignment = align_center
        cell.border = border_thin

    for row_idx, ano in enumerate(ANOS_EVALUACION, start=7):
        ind = indicadores[ano]
        p_tot = ind["provincial"]["Total"]
        p_fem = ind["provincial"]["Mujer"]
        p_masc = ind["provincial"]["Hombre"]
        terr = ind["metricas_territoriales"]

        ws1.cell(row=row_idx, column=1, value=ano).alignment = align_center
        ws1.cell(row=row_idx, column=2, value=ind["cohorte"]).alignment = align_center
        ws1.cell(row=row_idx, column=3, value=p_tot["num"]).number_format = "#,##0"
        ws1.cell(row=row_idx, column=4, value=p_tot["den"]).number_format = "#,##0"
        
        c_tot = ws1.cell(row=row_idx, column=5, value=p_tot["cob"] / 100.0)
        c_tot.number_format = "0.0%"
        c_tot.font = font_bold
        
        c_fem = ws1.cell(row=row_idx, column=6, value=p_fem["cob"] / 100.0)
        c_fem.number_format = "0.0%"
        
        c_masc = ws1.cell(row=row_idx, column=7, value=p_masc["cob"] / 100.0)
        c_masc.number_format = "0.0%"

        c_gpi = ws1.cell(row=row_idx, column=8, value=terr["gpi"] if terr["gpi"] else "-")
        c_gpi.alignment = align_center

        c_brecha = ws1.cell(row=row_idx, column=9, value=terr["brecha_max_min"] / 100.0)
        c_brecha.number_format = "0.0%"

        for c_i in range(1, 10):
            ws1.cell(row=row_idx, column=c_i).border = border_thin
            if row_idx % 2 == 0:
                ws1.cell(row=row_idx, column=c_i).fill = fill_zebra

    # Ficha Metodológica
    start_r = 21
    ws1.cell(row=start_r, column=1, value="2. FICHA TÉCNICA Y RESPALDO METODOLÓGICO").font = font_sec
    
    ficha_items = [
        ("Nombre del Indicador:", "Cobertura de vacunación contra Virus del Papiloma Humano (VPH) en personas de 15 años cumplidos."),
        ("Fórmula de Cálculo:", "Cobertura (%) = [ Numerador / Denominador ] × 100"),
        ("Definición de Numerador:", "Personas residentes en la comuna con esquema completo de VPH al cumplir 15 años (registradas en RNI)."),
        ("Definición de Denominador:", "Población objetivo = Nacidos vivos de la cohorte correspondiente (DEIS) menos Defunciones ocurridas entre 0 y 14 años (DEIS)."),
        ("Criterio de Residencia:", "Estricto por comuna de residencia del paciente informada en RNI y en estadísticas vitales DEIS (Comunas 10301 a 10307)."),
        ("Filtros MINSAL Aplicados:", "VACUNA_ADMINISTRADA == 'SI', REGISTRO_ELIMINADO != 'SI', Exclusión total de registros con valor 'EPRO'."),
        ("Transición de Esquema:", "2014-2022: Esquema 2 dosis Tetravalente en mujeres (hombres desde 2019). | 2023-2026: Dosis Única Nonavalente (Gardasil 9)."),
        ("Nota de Corte Año 2026:", "El año 2026 corresponde a un corte preliminar en curso de la campaña escolar (datos en desarrollo continuo)."),
        ("Fuentes Oficiales:", "Registro Nacional de Inmunizaciones (RNI) & Departamento de Estadísticas e Información de Salud (DEIS) — MINSAL Chile."),
    ]

    for idx, (label, val) in enumerate(ficha_items, start=start_r + 1):
        ws1.cell(row=idx, column=1, value=label).font = font_bold
        ws1.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=9)
        cell_v = ws1.cell(row=idx, column=2, value=val)
        cell_v.font = font_regular
        cell_v.alignment = align_left

    auto_fit_columns(ws1)

    # -------------------------------------------------------------------------
    # FUNCION GENERICA PARA HOJAS DEL INDICADOR (TOTAL, MUJERES, HOMBRES)
    # -------------------------------------------------------------------------
    def crear_hoja_indicador(wb, title, sexo_key, fill_header_color):
        ws = wb.create_sheet(title=title)
        ws.views.sheetView[0].showGridLines = True

        ws.merge_cells("A1:N2")
        ws["A1"] = f"INDICADOR VPH A LOS 15 AÑOS — {title.upper()}"
        ws["A1"].font = font_title
        ws["A1"].fill = fill_header_color
        ws["A1"].alignment = align_center

        ws.merge_cells("A3:N3")
        ws["A3"] = "SERIE HISTÓRICA 2015 - 2026 (COHORTES 2000 A 2011) — SERVICIO DE SALUD OSORNO"
        ws["A3"].font = font_subtitle
        ws["A3"].fill = fill_secondary
        ws["A3"].alignment = align_center

        # Cabeceras
        headers_sup = ["Comuna de Residencia", "Cód. Comuna"]
        for ano in ANOS_EVALUACION:
            headers_sup.append(f"{ano} (Coh. {COHORTES_MAP[ano]})")

        for col_idx, h in enumerate(headers_sup, start=1):
            c = ws.cell(row=5, column=col_idx, value=h)
            c.font = font_header
            c.fill = fill_header_color
            c.alignment = align_center
            c.border = border_thin

        # Filas por comuna
        for r_i, com_cod in enumerate(CODIGOS_COMUNAS, start=6):
            com_nom = COMUNAS_OSORNO[com_cod]
            ws.cell(row=r_i, column=1, value=com_nom).font = font_bold
            ws.cell(row=r_i, column=2, value=com_cod).alignment = align_center

            for c_i, ano in enumerate(ANOS_EVALUACION, start=3):
                val_cob = indicadores[ano]["comunas"][com_cod][sexo_key]["cob"]
                cell = ws.cell(row=r_i, column=c_i, value=val_cob / 100.0)
                cell.number_format = "0.0%"
                cell.alignment = align_right
                cell.border = border_thin
                if r_i % 2 == 0:
                    cell.fill = fill_zebra

            ws.cell(row=r_i, column=1).border = border_thin
            ws.cell(row=r_i, column=2).border = border_thin

        # Fila Total Provincial
        row_tot = 6 + len(CODIGOS_COMUNAS)
        ws.cell(row=row_tot, column=1, value="TOTAL S.S. OSORNO").font = font_bold
        ws.cell(row=row_tot, column=2, value="PROVINCIA").font = font_bold
        ws.cell(row=row_tot, column=2).alignment = align_center

        for c_i, ano in enumerate(ANOS_EVALUACION, start=3):
            val_cob = indicadores[ano]["provincial"][sexo_key]["cob"]
            cell = ws.cell(row=row_tot, column=c_i, value=val_cob / 100.0)
            cell.font = font_bold
            cell.number_format = "0.0%"
            cell.alignment = align_right
            cell.fill = fill_total_row
            cell.border = border_double_bottom

        ws.cell(row=row_tot, column=1).fill = fill_total_row
        ws.cell(row=row_tot, column=1).border = border_double_bottom
        ws.cell(row=row_tot, column=2).fill = fill_total_row
        ws.cell(row=row_tot, column=2).border = border_double_bottom

        # Tabla de Detalle de Numeradores
        row_det = row_tot + 3
        ws.cell(row=row_det, column=1, value="DETALLE DE NUMERADORES (PERSONAS CON ESQUEMA COMPLETO)").font = font_sec

        ws.cell(row=row_det + 1, column=1, value="Comuna").font = font_header
        ws.cell(row=row_det + 1, column=1).fill = fill_primary
        ws.cell(row=row_det + 1, column=2, value="Cód.").font = font_header
        ws.cell(row=row_det + 1, column=2).fill = fill_primary

        for c_i, ano in enumerate(ANOS_EVALUACION, start=3):
            c = ws.cell(row=row_det + 1, column=c_i, value=str(ano))
            c.font = font_header
            c.fill = fill_primary
            c.alignment = align_center

        for r_i, com_cod in enumerate(CODIGOS_COMUNAS, start=row_det + 2):
            ws.cell(row=r_i, column=1, value=COMUNAS_OSORNO[com_cod])
            ws.cell(row=r_i, column=2, value=com_cod).alignment = align_center
            for c_i, ano in enumerate(ANOS_EVALUACION, start=3):
                num = indicadores[ano]["comunas"][com_cod][sexo_key]["num"]
                cell = ws.cell(row=r_i, column=c_i, value=num)
                cell.number_format = "#,##0"
                cell.alignment = align_right
                cell.border = border_thin

        # Total provincial num
        r_tot_n = row_det + 2 + len(CODIGOS_COMUNAS)
        ws.cell(row=r_tot_n, column=1, value="TOTAL S.S. OSORNO").font = font_bold
        ws.cell(row=r_tot_n, column=2, value="PROVINCIA").font = font_bold
        for c_i, ano in enumerate(ANOS_EVALUACION, start=3):
            cell = ws.cell(row=r_tot_n, column=c_i, value=indicadores[ano]["provincial"][sexo_key]["num"])
            cell.font = font_bold
            cell.number_format = "#,##0"
            cell.alignment = align_right
            cell.fill = fill_total_row
            cell.border = border_double_bottom

        auto_fit_columns(ws)

    # Crear Hojas 2, 3 y 4
    crear_hoja_indicador(wb, "Indicador 15 Años - Total", "Total", fill_primary)
    crear_hoja_indicador(wb, "Indicador 15 Años - Mujeres", "Mujer", fill_fem)
    crear_hoja_indicador(wb, "Indicador 15 Años - Hombres", "Hombre", fill_masc)

    # -------------------------------------------------------------------------
    # HOJA 5: DOSIS ANUALES POR COMUNA Y SEXO (2014-2026)
    # -------------------------------------------------------------------------
    ws5 = wb.create_sheet(title="Dosis Anuales (2014-2026)")
    ws5.views.sheetView[0].showGridLines = True

    ws5.merge_cells("A1:J2")
    ws5["A1"] = "PRODUCCIÓN ANUAL DE DOSIS VPH (2014 - 2026) — SERVICIO DE SALUD OSORNO"
    ws5["A1"].font = font_title
    ws5["A1"].fill = fill_accent
    ws5["A1"].alignment = align_center

    headers_dosis = [
        "Año Administración", "Cód. Comuna", "Comuna Residencia", "Sexo",
        "1ª Dosis", "2ª Dosis", "3ª Dosis", "Dosis Única", "Otra Dosis", "Total Dosis"
    ]
    for col_idx, h in enumerate(headers_dosis, start=1):
        cell = ws5.cell(row=4, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_primary
        cell.alignment = align_center
        cell.border = border_thin

    curr_row = 5
    for ano_adm in sorted(PROGRAMATICAS_FILES.keys()):
        for com_cod in CODIGOS_COMUNAS:
            com_nom = COMUNAS_OSORNO[com_cod]
            for sexo in ("Mujer", "Hombre"):
                entry = dosis_anuales[ano_adm][com_cod][sexo]
                ws5.cell(row=curr_row, column=1, value=ano_adm).alignment = align_center
                ws5.cell(row=curr_row, column=2, value=com_cod).alignment = align_center
                ws5.cell(row=curr_row, column=3, value=com_nom)
                ws5.cell(row=curr_row, column=4, value=sexo).alignment = align_center
                ws5.cell(row=curr_row, column=5, value=entry["1ª Dosis"]).number_format = "#,##0"
                ws5.cell(row=curr_row, column=6, value=entry["2ª Dosis"]).number_format = "#,##0"
                ws5.cell(row=curr_row, column=7, value=entry["3ª Dosis"]).number_format = "#,##0"
                ws5.cell(row=curr_row, column=8, value=entry["Dosis Única"]).number_format = "#,##0"
                ws5.cell(row=curr_row, column=9, value=entry["Otra Dosis"]).number_format = "#,##0"
                ws5.cell(row=curr_row, column=10, value=entry["Total"]).number_format = "#,##0"

                for c_i in range(1, 11):
                    ws5.cell(row=curr_row, column=c_i).border = border_thin
                curr_row += 1

    auto_fit_columns(ws5)

    # -------------------------------------------------------------------------
    # HOJA 6: DENOMINADORES DEIS (NACIMIENTOS Y DEFUNCIONES)
    # -------------------------------------------------------------------------
    ws6 = wb.create_sheet(title="Denominadores DEIS")
    ws6.views.sheetView[0].showGridLines = True

    ws6.merge_cells("A1:O2")
    ws6["A1"] = "DENOMINADORES DEIS: NACIDOS VIVOS MENOS DEFUNCIONES (<15 AÑOS)"
    ws6["A1"].font = font_title
    ws6["A1"].fill = fill_primary
    ws6["A1"].alignment = align_center

    ws6.cell(row=4, column=1, value="Comuna").font = font_header
    ws6.cell(row=4, column=1).fill = fill_primary
    ws6.cell(row=4, column=2, value="Cód.").font = font_header
    ws6.cell(row=4, column=2).fill = fill_primary
    ws6.cell(row=4, column=3, value="Componente").font = font_header
    ws6.cell(row=4, column=3).fill = fill_primary

    for c_i, ano in enumerate(ANOS_EVALUACION, start=4):
        c = ws6.cell(row=4, column=c_i, value=f"{ano} (Coh. {COHORTES_MAP[ano]})")
        c.font = font_header
        c.fill = fill_primary
        c.alignment = align_center

    curr_row = 5
    for com_cod in CODIGOS_COMUNAS:
        com_nom = COMUNAS_OSORNO[com_cod]
        
        # Fila Nacimientos
        ws6.cell(row=curr_row, column=1, value=com_nom).font = font_bold
        ws6.cell(row=curr_row, column=2, value=com_cod).alignment = align_center
        ws6.cell(row=curr_row, column=3, value="Nacidos Vivos (DEIS)")
        for c_i, ano in enumerate(ANOS_EVALUACION, start=4):
            val = indicadores[ano]["comunas"][com_cod]["Total"]["nac"]
            ws6.cell(row=curr_row, column=c_i, value=val).number_format = "#,##0"
            ws6.cell(row=curr_row, column=c_i).border = border_thin
        curr_row += 1

        # Fila Defunciones
        ws6.cell(row=curr_row, column=1, value=com_nom)
        ws6.cell(row=curr_row, column=2, value=com_cod).alignment = align_center
        ws6.cell(row=curr_row, column=3, value="Defunciones <15 años (DEIS)")
        for c_i, ano in enumerate(ANOS_EVALUACION, start=4):
            val = indicadores[ano]["comunas"][com_cod]["Total"]["def"]
            ws6.cell(row=curr_row, column=c_i, value=val).number_format = "#,##0"
            ws6.cell(row=curr_row, column=c_i).border = border_thin
        curr_row += 1

        # Fila Denominador Neto
        ws6.cell(row=curr_row, column=1, value=com_nom).font = font_bold
        ws6.cell(row=curr_row, column=2, value=com_cod).alignment = align_center
        ws6.cell(row=curr_row, column=3, value="Población Objetivo Neta").font = font_bold
        for c_i, ano in enumerate(ANOS_EVALUACION, start=4):
            val = indicadores[ano]["comunas"][com_cod]["Total"]["den"]
            cell = ws6.cell(row=curr_row, column=c_i, value=val)
            cell.font = font_bold
            cell.number_format = "#,##0"
            cell.fill = fill_zebra
            cell.border = border_thin
        curr_row += 2

    auto_fit_columns(ws6)

    # -------------------------------------------------------------------------
    # HOJA 7: DETALLE POR ESTABLECIMIENTO DE SALUD
    # -------------------------------------------------------------------------
    ws7 = wb.create_sheet(title="Establecimientos")
    ws7.views.sheetView[0].showGridLines = True

    ws7.merge_cells("A1:F2")
    ws7["A1"] = "DOSIS VPH ADMINISTRADAS POR ESTABLECIMIENTO DE SALUD (2014 - 2026)"
    ws7["A1"].font = font_title
    ws7["A1"].fill = fill_primary
    ws7["A1"].alignment = align_center

    headers_estab = ["Establecimiento de Salud", "VPH Tetravalente", "VPH Nonavalente", "VPH Bivalente", "Otras / No Espec.", "Total Dosis"]
    for col_idx, h in enumerate(headers_estab, start=1):
        cell = ws7.cell(row=4, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_primary
        cell.alignment = align_center
        cell.border = border_thin

    row_est = 5
    for est, vacs in sorted(estab_detalle.items(), key=lambda x: sum(x[1].values()), reverse=True):
        ws7.cell(row=row_est, column=1, value=est).font = font_bold
        t_tetra = vacs.get("VPH Tetravalente (Gardasil 4)", 0)
        t_nona = vacs.get("VPH Nonavalente (Gardasil 9)", 0)
        t_biv = vacs.get("VPH Bivalente (Cervarix)", 0)
        t_otra = vacs.get("VPH Otra / No Especificada", 0)
        tot_e = sum(vacs.values())

        ws7.cell(row=row_est, column=2, value=t_tetra).number_format = "#,##0"
        ws7.cell(row=row_est, column=3, value=t_nona).number_format = "#,##0"
        ws7.cell(row=row_est, column=4, value=t_biv).number_format = "#,##0"
        ws7.cell(row=row_est, column=5, value=t_otra).number_format = "#,##0"
        ws7.cell(row=row_est, column=6, value=tot_e).number_format = "#,##0"

        for c_i in range(1, 7):
            ws7.cell(row=row_est, column=c_i).border = border_thin
            if row_est % 2 == 0:
                ws7.cell(row=row_est, column=c_i).fill = fill_zebra
        row_est += 1

    auto_fit_columns(ws7)

    # Guardar Excel Maestro
    wb.save(OUTPUT_EXCEL)
    print(f"\n  [OK] Reporte Maestro Excel guardado en: {OUTPUT_EXCEL}")


# =============================================================================
# EJECUCIÓN PRINCIPAL
# =============================================================================

def main():
    t_start = time.time()
    print("=" * 75)
    print("OBSERVATORIO EPIDEMIOLÓGICO VPH (2014-2026) — SERVICIO DE SALUD OSORNO")
    print(f"Inicio de Ejecución: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 75)

    # 1. Leer RNI Programáticas (2014-2026)
    personas, dosis_anuales, estab_detalle, vacunas_por_ano, estab_matriz = leer_programaticas_vph()

    # 2. Evaluar Esquemas y Cohortes
    vacunados = procesar_cohortes_vacunadas(personas)

    # 3. Leer Nacimientos DEIS (2000-2021)
    nacimientos = leer_nacimientos_por_cohorte()

    # 4. Leer Defunciones DEIS (1999-2026)
    defunciones = leer_defunciones_por_cohorte()

    # 5. Calcular Indicadores y Métricas Epidemiológicas
    indicadores = calcular_indicadores_completos(vacunados, nacimientos, defunciones, dosis_anuales)

    # 6. Exportar para Dashboard Web (JSON y JS)
    exportar_datos_dashboard(indicadores, dosis_anuales, estab_detalle, vacunas_por_ano, estab_matriz)

    # 7. Generar Reporte Maestro Excel
    crear_reporte_excel_maestro(indicadores, dosis_anuales, estab_detalle, vacunas_por_ano)

    t_total = time.time() - t_start
    print("\n" + "=" * 75)
    print(f"[OK] PROCESO FINALIZADO EXITOSAMENTE EN {t_total:.1f} SEGUNDOS ({t_total/60:.1f} min)")
    print("=" * 75)


if __name__ == "__main__":
    main()
