#!/usr/bin/env python
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
# -*- coding: utf-8 -*-
"""
===============================================================================
CÁLCULO DEL INDICADOR:
  "Porcentaje de población de 15 años con esquema completo de vacunación de VPH"

Estrategia Nacional de Cáncer - Objetivo Estratégico 7
Serie Histórica Completa (2019-2026) por Comuna de Residencia y Sexo

Fórmula:
  (Población de 15 años con esquema completo VPH / Nº personas de 15 años) × 100

Criterios de esquema completo:
  - 2019-2024 (Cohortes 2004-2009): 2+ dosis registradas (4° + 5° Básico) o dosis Única
  - 2025-2026 (Cohortes 2010-2011): 1+ dosis registrada (dosis Única, nuevo calendario Nonavalente)

Incluye todas las vacunas VPH (Tetravalente, Nonavalente, Bivalente).
Desagregación por Sexo (Mujeres, Hombres, Total) y por Comuna de Residencia.

Fuentes:
  - Numerador: Registro Nacional de Inmunizaciones (RNI) — Bases Programáticas por Residencia
  - Denominador: Nacidos Vivos DEIS (por residencia) - Defunciones 0-14 años DEIS (por residencia)

Nota técnica: 2026 es el único año incompleto (en desarrollo, corte al 31/07/2026).
===============================================================================
"""

import csv
import os
import sys
import time
from collections import defaultdict
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: Se requiere openpyxl. Instale con: pip install openpyxl")
    sys.exit(1)

# =============================================================================
# CONFIGURACIÓN
# =============================================================================

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
BD_MINSAL_DIR = os.path.join(os.path.dirname(BASE_DIR), "BASE DATOS MINSAL")

def get_db_dir(year):
    """Devuelve la ruta al directorio de la base de datos según el año"""
    if int(year) <= 2024:
        return os.path.join(BD_MINSAL_DIR, "2000-2024")
    else:
        return os.path.join(BD_MINSAL_DIR, str(year))

# Años de medición del indicador (cumplen 15 años)
ANOS_REPORTE = [2019, 2020, 2021, 2022, 2023, 2024, 2025, 2026]
COHORTES_NACIMIENTO = {ano: ano - 15 for ano in ANOS_REPORTE}

# Archivos de Programáticas disponibles (2015 a 2026)
PROGRAMATICAS_FILES = {
    2015: "Programáticas_Residencia_2015",
    2016: "Programáticas_Residencia_2016",
    2017: "Programáticas_Residencia_2017",
    2018: "Programáticas_Residencia_2018",
    2019: "Programáticas_Residencia_2019.csv",
    2020: "Programáticas_Residencia_2020.csv",
    2021: "Programáticas_Residencia_2021.csv",
    2022: "Programáticas_Residencia_2022.csv",
    2023: "Programáticas_Residencia_2023.csv",
    2024: "Programáticas_Residencia_2024.csv",
    2025: "Programáticas_Residencia_2025.csv",
    2026: "Programáticas_Residencia_2026.csv",
}

# Comunas Servicio de Salud Osorno (Código 23)
COMUNAS_OSORNO = {
    "10301": "Osorno",
    "10302": "Puerto Octay",
    "10303": "Purranque",
    "10304": "Puyehue",
    "10305": "Río Negro",
    "10306": "San Juan de la Costa",
    "10307": "San Pablo",
}

CODIGOS_COMUNAS = sorted(COMUNAS_OSORNO.keys())

VACUNAS_VPH_KEYWORDS = ["VPH", "PAPILOMA", "GARDASIL", "CERVARIX"]


def es_vacuna_vph(nombre_vacuna):
    if not nombre_vacuna:
        return False
    nombre_upper = str(nombre_vacuna).upper()
    return any(kw in nombre_upper for kw in VACUNAS_VPH_KEYWORDS)


def clasificar_tipo_dosis(dosis_str):
    if not dosis_str:
        return "Otra"
    d = str(dosis_str).strip().upper()
    if "ÚNICA" in d or "UNICA" in d:
        return "Dosis Única"
    if "1" in d:
        return "1ª Dosis"
    if "2" in d:
        return "2ª Dosis"
    if "3" in d:
        return "3ª Dosis"
    return "Otra Dosis"


def normalizar_sexo(val):
    if val is None:
        return "Otro"
    v = str(val).strip().upper()
    if v in ("1", "1.0", "H", "HOMBRE", "HOMBRES", "M", "MASCULINO"):
        return "Hombre"
    if v in ("2", "2.0", "M", "MUJER", "MUJERES", "F", "FEMENINO"):
        return "Mujer"
    return "Otro"


def parsear_fecha_nacimiento(fecha_str):
    if not fecha_str:
        return None
    fecha_str = str(fecha_str).strip()
    if " " in fecha_str:
        fecha_str = fecha_str.split(" ")[0]
    formatos = ["%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%d.%m.%Y", "%Y%m%d"]
    for fmt in formatos:
        try:
            return datetime.strptime(fecha_str, fmt).year
        except ValueError:
            pass
    if len(fecha_str) >= 4 and fecha_str[:4].isdigit():
        y = int(fecha_str[:4])
        if 1990 <= y <= 2030:
            return y
    return None


def leer_programaticas_vph():
    """
    Lee todas las bases programáticas (2015-2026) por residencia.
    Retorna:
      - registros_personas: dict[run] -> info persona (nacimiento, comuna, sexo, dosis, etc.)
      - dosis_anuales: dict[ano_adm][comuna][sexo][tipo_dosis] -> count
      - detalle_estab: dict[(ano_medicion, cod_comuna, nombre_estab)] -> set(runs)
    """
    print("=" * 70)
    print("PASO 1: Lectura de Bases Programáticas RNI (Criterio Residencia)")
    print("=" * 70)

    registros_personas = defaultdict(lambda: {
        "dosis": [],
        "comuna_residencia": None,
        "ano_nacimiento": None,
        "sexo": None,
        "estab_ultimo": "Desconocido",
    })

    # dosis_anuales[ano_adm][cod_comuna][sexo][tipo_dosis] -> int
    dosis_anuales = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(int))))
    detalle_estab = defaultdict(lambda: defaultdict(lambda: defaultdict(set)))

    anos_nac_interes = set(COHORTES_NACIMIENTO.values())

    for ano, filename in sorted(PROGRAMATICAS_FILES.items()):
        filepath = os.path.join(get_db_dir(ano), filename)
        if not os.path.exists(filepath):
            print(f"  [AVISO] Archivo no encontrado: {filename}")
            continue

        print(f"  Leyendo año {ano} ({filename})...")
        t0 = time.time()
        vph_validos = 0

        with open(filepath, "r", encoding="latin-1", errors="replace") as f:
            first_line = f.readline()
            delimiter = "|" if "|" in first_line else (";" if ";" in first_line else ",")
            f.seek(0)
            reader = csv.reader(f, delimiter=delimiter)
            
            header = next(reader)
            col_map = {col.strip().upper(): i for i, col in enumerate(header)}

            idx_vac = col_map.get("NOMBRE_VACUNA")
            idx_dosis = col_map.get("DOSIS")
            idx_adm = col_map.get("VACUNA_ADMINISTRADA")
            idx_elim = col_map.get("REGISTRO_ELIMINADO")
            idx_crit = col_map.get("CRITERIO_ELEGIBILIDAD")
            idx_run = col_map.get("RUN")
            idx_fnac = col_map.get("FECHA_NACIMIENTO")
            idx_sexo = col_map.get("SEXO")
            idx_com_res = col_map.get("COD_COMUNA_RESID")
            idx_estab = col_map.get("ESTABLECIMIENTO")

            if idx_vac is None or idx_run is None:
                print(f"    [ERROR] Columnas clave faltantes en {filename}")
                continue

            for row in reader:
                if len(row) <= max(idx_vac, idx_run):
                    continue

                # Filtro: Vacuna VPH
                if not es_vacuna_vph(row[idx_vac]):
                    continue

                # Filtros obligatorios de calidad MINSAL
                if idx_adm is not None and len(row) > idx_adm:
                    if row[idx_adm].strip().upper() != "SI":
                        continue
                if idx_elim is not None and len(row) > idx_elim:
                    if row[idx_elim].strip().upper() == "SI":
                        continue
                if idx_crit is not None and len(row) > idx_crit:
                    if row[idx_crit].strip().upper() == "EPRO":
                        continue
                if idx_dosis is not None and len(row) > idx_dosis:
                    if row[idx_dosis].strip().upper() == "EPRO":
                        continue

                # Filtro: Comuna de residencia en S.S. Osorno
                com_res = ""
                if idx_com_res is not None and len(row) > idx_com_res:
                    com_res = row[idx_com_res].strip()
                    if com_res.isdigit():
                        com_res = str(int(com_res))

                if com_res not in COMUNAS_OSORNO:
                    continue

                # Extraer datos
                dosis_raw = row[idx_dosis].strip() if idx_dosis is not None and len(row) > idx_dosis else ""
                tipo_dosis = clasificar_tipo_dosis(dosis_raw)
                
                sexo_raw = row[idx_sexo].strip() if idx_sexo is not None and len(row) > idx_sexo else ""
                sexo = normalizar_sexo(sexo_raw)

                estab = row[idx_estab].strip() if idx_estab is not None and len(row) > idx_estab else "Desconocido"
                run = row[idx_run].strip()

                # Registro de dosis anuales
                dosis_anuales[ano][com_res][sexo][tipo_dosis] += 1
                dosis_anuales[ano][com_res][sexo]["Total"] += 1
                dosis_anuales[ano]["TOTAL"][sexo][tipo_dosis] += 1
                dosis_anuales[ano]["TOTAL"][sexo]["Total"] += 1
                dosis_anuales[ano][com_res]["Total"][tipo_dosis] += 1
                dosis_anuales[ano][com_res]["Total"]["Total"] += 1
                dosis_anuales[ano]["TOTAL"]["Total"][tipo_dosis] += 1
                dosis_anuales[ano]["TOTAL"]["Total"]["Total"] += 1

                # Extracción para cohorte (si tiene RUN y año de nacimiento)
                if run and run not in ("", "0"):
                    fnac_raw = row[idx_fnac].strip() if idx_fnac is not None and len(row) > idx_fnac else ""
                    ano_nac = parsear_fecha_nacimiento(fnac_raw)

                    p = registros_personas[run]
                    p["dosis"].append({
                        "ano_vacunacion": ano,
                        "dosis": dosis_raw,
                        "tipo_dosis": tipo_dosis,
                        "vacuna": row[idx_vac].strip(),
                    })
                    if ano_nac and ano_nac in anos_nac_interes:
                        p["ano_nacimiento"] = ano_nac
                    if com_res:
                        p["comuna_residencia"] = com_res
                    if sexo != "Otro":
                        p["sexo"] = sexo
                    if estab and estab != "Desconocido":
                        p["estab_ultimo"] = estab

                vph_validos += 1

        print(f"    -> {vph_validos:,} registros VPH válidos procesados ({time.time()-t0:.1f}s)")

    print(f"\n  Total de personas únicas con registro VPH: {len(registros_personas):,}")
    return registros_personas, dosis_anuales


def determinar_esquema_completo(registros_personas):
    """
    Evalúa si cada persona cumple esquema completo según su cohorte y reglas históricas.
    """
    print("\n" + "=" * 70)
    print("PASO 2: Evaluación de Esquema Completo por Persona y Sexo")
    print("=" * 70)

    # resultado[ano_medicion][cod_comuna][sexo] -> set(runs)
    completos = defaultdict(lambda: defaultdict(lambda: defaultdict(set)))
    estab_detalle = defaultdict(int)

    # Mapeo inverso de cohorte a año de medición
    cohorte_a_medicion = {v: k for k, v in COHORTES_NACIMIENTO.items()}

    for run, p in registros_personas.items():
        ano_nac = p["ano_nacimiento"]
        if ano_nac not in cohorte_a_medicion:
            continue

        ano_medicion = cohorte_a_medicion[ano_nac]
        comuna = p["comuna_residencia"]
        if not comuna or comuna not in COMUNAS_OSORNO:
            continue

        sexo = p["sexo"] if p["sexo"] in ("Hombre", "Mujer") else "Mujer"  # default si no especificado
        num_dosis = len(p["dosis"])
        tiene_dosis_unica = any(d["tipo_dosis"] == "Dosis Única" for d in p["dosis"])

        # Regla según cohorte
        if ano_medicion <= 2024:
            es_completo = (num_dosis >= 2) or tiene_dosis_unica
        else:
            es_completo = (num_dosis >= 1)

        if es_completo:
            completos[ano_medicion][comuna][sexo].add(run)
            completos[ano_medicion][comuna]["Total"].add(run)
            completos[ano_medicion]["TOTAL"][sexo].add(run)
            completos[ano_medicion]["TOTAL"]["Total"].add(run)

            estab = p.get("estab_ultimo", "Desconocido")
            estab_detalle[(ano_medicion, comuna, estab)] += 1

    for ano in ANOS_REPORTE:
        c = COHORTES_NACIMIENTO[ano]
        tot = len(completos[ano]["TOTAL"]["Total"])
        fem = len(completos[ano]["TOTAL"]["Mujer"])
        masc = len(completos[ano]["TOTAL"]["Hombre"])
        print(f"  Año {ano} (Cohorte {c}): Total={tot:,} (Mujeres={fem:,}, Hombres={masc:,})")

    return completos, estab_detalle


def leer_nacimientos_por_cohorte():
    """
    Lee archivos NAC2004 a NAC2021 de DEIS por residencia.
    Retorna: nacimientos[ano_nac][comuna][sexo] -> count
    """
    print("\n" + "=" * 70)
    print("PASO 3: Lectura de Nacidos Vivos DEIS (Criterio Residencia y Sexo)")
    print("=" * 70)

    nacimientos = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
    anos_necesarios = sorted(set(COHORTES_NACIMIENTO.values()))

    for ano_nac in anos_necesarios:
        f_xlsx = os.path.join(get_db_dir(ano_nac), f"NAC{ano_nac}.xlsx")
        f_csv = os.path.join(get_db_dir(ano_nac), f"NAC{ano_nac}.csv")

        if os.path.exists(f_xlsx):
            print(f"  Leyendo nacimientos {ano_nac} (Excel)...")
            wb = openpyxl.load_workbook(f_xlsx, read_only=True, data_only=True)
            ws = wb.active
            header = [str(c.value).strip().upper() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

            idx_com = header.index("COMUNA") if "COMUNA" in header else (header.index("COD_COMUNA") if "COD_COMUNA" in header else None)
            idx_sexo = header.index("SEXO") if "SEXO" in header else next((i for i, h in enumerate(header) if "SEX" in h), None)

            if idx_com is not None:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if len(row) > idx_com and row[idx_com] is not None:
                        try:
                            val = int(row[idx_com]) if isinstance(row[idx_com], (int, float)) else int(str(row[idx_com]).strip())
                            com_str = str(val)
                        except (ValueError, TypeError):
                            continue
                        if com_str in COMUNAS_OSORNO:
                            sexo = "Otro"
                            if idx_sexo is not None and len(row) > idx_sexo and row[idx_sexo] is not None:
                                sexo = normalizar_sexo(row[idx_sexo])
                            if sexo not in ("Hombre", "Mujer"):
                                sexo = "Mujer"

                            nacimientos[ano_nac][com_str][sexo] += 1
                            nacimientos[ano_nac][com_str]["Total"] += 1
                            nacimientos[ano_nac]["TOTAL"][sexo] += 1
                            nacimientos[ano_nac]["TOTAL"]["Total"] += 1
            wb.close()

        elif os.path.exists(f_csv):
            print(f"  Leyendo nacimientos {ano_nac} (CSV)...")
            with open(f_csv, "r", encoding="latin-1", errors="replace") as f:
                first = f.readline()
                delim = ";" if ";" in first else ("," if "," in first else "|")
                f.seek(0)
                reader = csv.reader(f, delimiter=delim)
                header = [h.strip().upper() for h in next(reader)]

                idx_com = header.index("COMUNA") if "COMUNA" in header else (header.index("COD_COMUNA") if "COD_COMUNA" in header else None)
                idx_sexo = header.index("SEXO") if "SEXO" in header else next((i for i, h in enumerate(header) if "SEX" in h), None)

                if idx_com is not None:
                    for row in reader:
                        if len(row) > idx_com:
                            val = row[idx_com].strip()
                            if val.isdigit():
                                com_str = str(int(val))
                                if com_str in COMUNAS_OSORNO:
                                    sexo = "Otro"
                                    if idx_sexo is not None and len(row) > idx_sexo:
                                        sexo = normalizar_sexo(row[idx_sexo].strip())
                                    if sexo not in ("Hombre", "Mujer"):
                                        sexo = "Mujer"

                                    nacimientos[ano_nac][com_str][sexo] += 1
                                    nacimientos[ano_nac][com_str]["Total"] += 1
                                    nacimientos[ano_nac]["TOTAL"][sexo] += 1
                                    nacimientos[ano_nac]["TOTAL"]["Total"] += 1

        tot = nacimientos[ano_nac]["TOTAL"]["Total"]
        fem = nacimientos[ano_nac]["TOTAL"]["Mujer"]
        masc = nacimientos[ano_nac]["TOTAL"]["Hombre"]
        print(f"    -> Cohorte {ano_nac}: Total={tot:,} (Mujeres={fem:,}, Hombres={masc:,})")

    return nacimientos


def leer_defunciones_por_cohorte():
    """
    Lee archivos DEF2006 a DEF2026 de DEIS y acumula defunciones 0-14 años por cohorte y sexo.
    """
    print("\n" + "=" * 70)
    print("PASO 4: Lectura de Defunciones DEIS 0-14 años (Criterio Residencia y Sexo)")
    print("=" * 70)

    defunciones = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
    anos_cohortes = set(COHORTES_NACIMIENTO.values())

    anos_def_necesarios = set()
    for c in anos_cohortes:
        for d in range(c, c + 15):
            anos_def_necesarios.add(d)

    for ano_def in sorted(anos_def_necesarios):
        f_xlsx = os.path.join(get_db_dir(ano_def), f"DEF{ano_def}.xlsx")
        f_csv = os.path.join(get_db_dir(ano_def), f"DEF{ano_def}.csv")

        if os.path.exists(f_xlsx):
            print(f"  Leyendo defunciones {ano_def} (Excel)...")
            wb = openpyxl.load_workbook(f_xlsx, read_only=True, data_only=True)
            ws = wb.active
            header = [str(c.value).strip().upper() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

            if "ANO1_NAC" in header and "ANO2_NAC" in header and "EDAD_TIPO" in header and "EDAD_CANT" in header:
                idx_a1 = header.index("ANO1_NAC")
                idx_a2 = header.index("ANO2_NAC")
                idx_t = header.index("EDAD_TIPO")
                idx_c = header.index("EDAD_CANT")
                idx_s = next((i for i, h in enumerate(header) if "SEX" in h), None)
                idx_com = next((i for i, h in enumerate(header) if h in ("COMUNA", "RES_COMUNA", "COD_COMUNA")), None)

                for row in ws.iter_rows(min_row=2, values_only=True):
                    if len(row) > max(idx_a1, idx_a2, idx_t, idx_c):
                        try:
                            a1 = int(row[idx_a1]) if row[idx_a1] is not None else 0
                            a2 = int(row[idx_a2]) if row[idx_a2] is not None else 0
                            anac = a1 * 100 + a2
                        except (ValueError, TypeError):
                            continue

                        if anac not in anos_cohortes:
                            continue

                        try:
                            t = int(row[idx_t]) if row[idx_t] is not None else 0
                            c = int(row[idx_c]) if row[idx_c] is not None else 0
                        except (ValueError, TypeError):
                            continue

                        es_menor_15 = (t == 1 and c < 15) or (t in (2, 3, 4, 5))
                        if not es_menor_15:
                            continue

                        com_str = ""
                        if idx_com is not None and len(row) > idx_com and row[idx_com] is not None:
                            try:
                                com_str = str(int(row[idx_com])) if isinstance(row[idx_com], (int, float)) else str(row[idx_com]).strip()
                            except: pass

                        if com_str in COMUNAS_OSORNO:
                            sexo = "Mujer"
                            if idx_s is not None and len(row) > idx_s and row[idx_s] is not None:
                                sexo = normalizar_sexo(row[idx_s])
                            if sexo not in ("Hombre", "Mujer"):
                                sexo = "Mujer"

                            defunciones[anac][com_str][sexo] += 1
                            defunciones[anac][com_str]["Total"] += 1
                            defunciones[anac]["TOTAL"][sexo] += 1
                            defunciones[anac]["TOTAL"]["Total"] += 1
            wb.close()

        elif os.path.exists(f_csv):
            print(f"  Leyendo defunciones {ano_def} (CSV)...")
            with open(f_csv, "r", encoding="latin-1", errors="replace") as f:
                first = f.readline()
                delim = ";" if ";" in first else ("," if "," in first else "|")
                f.seek(0)
                reader = csv.reader(f, delimiter=delim)
                header = [h.strip().upper() for h in next(reader)]

                if "ANO1_NAC" in header and "ANO2_NAC" in header and "EDAD_TIPO" in header and "EDAD_CANT" in header:
                    idx_a1 = header.index("ANO1_NAC")
                    idx_a2 = header.index("ANO2_NAC")
                    idx_t = header.index("EDAD_TIPO")
                    idx_c = header.index("EDAD_CANT")
                    idx_s = next((i for i, h in enumerate(header) if "SEX" in h), None)
                    idx_com = next((i for i, h in enumerate(header) if h in ("COMUNA", "RES_COMUNA", "COD_COMUNA")), None)

                    for row in reader:
                        if len(row) > max(idx_a1, idx_a2, idx_t, idx_c):
                            try:
                                a1 = int(row[idx_a1].strip())
                                a2 = int(row[idx_a2].strip())
                                anac = a1 * 100 + a2
                            except: continue

                            if anac not in anos_cohortes:
                                continue

                            try:
                                t = int(row[idx_t].strip())
                                c = int(row[idx_c].strip())
                            except: continue

                            es_menor_15 = (t == 1 and c < 15) or (t in (2, 3, 4, 5))
                            if not es_menor_15:
                                continue

                            com_str = ""
                            if idx_com is not None and len(row) > idx_com:
                                val = row[idx_com].strip()
                                if val.isdigit():
                                    com_str = str(int(val))

                            if com_str in COMUNAS_OSORNO:
                                sexo = "Mujer"
                                if idx_s is not None and len(row) > idx_s:
                                    sexo = normalizar_sexo(row[idx_s].strip())
                                if sexo not in ("Hombre", "Mujer"):
                                    sexo = "Mujer"

                                defunciones[anac][com_str][sexo] += 1
                                defunciones[anac][com_str]["Total"] += 1
                                defunciones[anac]["TOTAL"][sexo] += 1
                                defunciones[anac]["TOTAL"]["Total"] += 1

    for c in sorted(anos_cohortes):
        tot = defunciones[c]["TOTAL"]["Total"]
        fem = defunciones[c]["TOTAL"]["Mujer"]
        masc = defunciones[c]["TOTAL"]["Hombre"]
        print(f"  Defunciones 0-14 años cohorte {c}: Total={tot} (Mujeres={fem}, Hombres={masc})")

    return defunciones


def calcular_indicadores_completos(completos, nacimientos, defunciones):
    """
    Calcula Numerador, Denominador y Cobertura para Total, Mujeres y Hombres por comuna y año.
    """
    print("\n" + "=" * 70)
    print("PASO 5: Consolidación y Cálculo del Indicador")
    print("=" * 70)

    resultados = {}

    for ano in ANOS_REPORTE:
        c = COHORTES_NACIMIENTO[ano]
        resultados[ano] = {}

        for cod_com in CODIGOS_COMUNAS + ["TOTAL"]:
            nom_com = COMUNAS_OSORNO.get(cod_com, "TOTAL S.S. OSORNO")
            resultados[ano][cod_com] = {
                "nombre": nom_com,
                "Total": {},
                "Mujer": {},
                "Hombre": {},
            }

            for sexo in ["Total", "Mujer", "Hombre"]:
                num = len(completos[ano][cod_com][sexo])
                nac = nacimientos[c][cod_com][sexo]
                defs = defunciones[c][cod_com][sexo]
                den = max(0, nac - defs)
                pct = (num / den * 100.0) if den > 0 else 0.0

                resultados[ano][cod_com][sexo] = {
                    "num": num,
                    "nac": nac,
                    "def": defs,
                    "den": den,
                    "pct": pct,
                }

        tot_res = resultados[ano]["TOTAL"]["Total"]
        fem_res = resultados[ano]["TOTAL"]["Mujer"]
        mas_res = resultados[ano]["TOTAL"]["Hombre"]
        print(f"\nAÑO {ano} (Cohorte {c}):")
        print(f"  TOTAL S.S. OSORNO: Num={tot_res['num']:,}, Den={tot_res['den']:,} -> {tot_res['pct']:.2f}%")
        print(f"  - Mujeres:         Num={fem_res['num']:,}, Den={fem_res['den']:,} -> {fem_res['pct']:.2f}%")
        print(f"  - Hombres:         Num={mas_res['num']:,}, Den={mas_res['den']:,} -> {mas_res['pct']:.2f}%")

    return resultados


def crear_reporte_excel_completo(resultados, dosis_anuales, estab_detalle, output_path):
    """
    Genera el libro Excel profesional con 7 hojas completas.
    """
    print("\n" + "=" * 70)
    print("PASO 6: Generación de Reporte Excel Integral (7 Hojas)")
    print("=" * 70)

    wb = openpyxl.Workbook()
    # Eliminar hoja inicial por defecto
    if wb.active is not None:
        wb.remove(wb.active)

    # Paleta de estilos corporativos DEIS / MINSAL
    color_primary = "1B365D"     # Azul marino institucional
    color_secondary = "2C5282"   # Azul intermedio
    color_header = "3182CE"      # Azul claro encabezado
    color_zebra = "F7FAFC"       # Gris muy claro
    color_accent = "DD6B20"      # Naranja meta/destacado
    color_note = "FFFFCC"        # Amarillo suave para notas
    color_fem = "D53F8C"         # Rosa/magenta institucional para mujeres
    color_masc = "2B6CB0"        # Azul para hombres

    font_title = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    font_subtitle = Font(name="Calibri", size=10, italic=True, color="E2E8F0")
    font_section = Font(name="Calibri", size=11, bold=True, color="1A202C")
    font_tbl_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    font_data = Font(name="Calibri", size=10)
    font_bold = Font(name="Calibri", size=10, bold=True)
    font_note = Font(name="Calibri", size=9, italic=True, color="4A5568")

    fill_primary = PatternFill(start_color=color_primary, end_color=color_primary, fill_type="solid")
    fill_secondary = PatternFill(start_color=color_secondary, end_color=color_secondary, fill_type="solid")
    fill_header = PatternFill(start_color=color_header, end_color=color_header, fill_type="solid")
    fill_fem = PatternFill(start_color=color_fem, end_color=color_fem, fill_type="solid")
    fill_masc = PatternFill(start_color=color_masc, end_color=color_masc, fill_type="solid")
    fill_zebra = PatternFill(start_color=color_zebra, end_color=color_zebra, fill_type="solid")
    fill_total = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    fill_note = PatternFill(start_color=color_note, end_color=color_note, fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color="CBD5E0"),
        right=Side(style="thin", color="CBD5E0"),
        top=Side(style="thin", color="CBD5E0"),
        bottom=Side(style="thin", color="CBD5E0"),
    )
    total_border = Border(
        left=Side(style="thin", color="CBD5E0"),
        right=Side(style="thin", color="CBD5E0"),
        top=Side(style="medium", color="1B365D"),
        bottom=Side(style="double", color="1B365D"),
    )

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # =========================================================================
    # HOJA 1: FICHA Y EVOLUCIÓN VPH
    # =========================================================================
    ws1 = wb.create_sheet(title="Ficha y Evolución VPH")
    ws1.views.sheetView[0].showGridLines = True

    ws1.merge_cells("A1:G1")
    ws1["A1"] = "FICHA TÉCNICA E HISTORIAL NORMATIVO DE VACUNACIÓN VPH (15 AÑOS) — S.S. OSORNO"
    ws1["A1"].font = font_title
    ws1["A1"].fill = fill_primary
    ws1["A1"].alignment = align_center
    ws1.row_dimensions[1].height = 32

    ws1.merge_cells("A2:G2")
    ws1["A2"] = "CRITERIO DE RESIDENCIA: Toda la información ha sido construida exclusivamente en base a la Comuna y Servicio de Salud de Residencia"
    ws1["A2"].font = font_subtitle
    ws1["A2"].fill = fill_secondary
    ws1["A2"].alignment = align_center
    ws1.row_dimensions[2].height = 20

    ficha_items = [
        ("Indicador", "Porcentaje de población de 15 años con esquema completo de vacunación contra VPH"),
        ("Fórmula General", "(Población de 15 años con esquema completo VPH / Nº personas estimadas de 15 años) × 100"),
        ("Criterio Territorial", "Servicio de Salud Osorno (Código 23) y sus 7 comunas por RESIDENCIA habitual"),
        ("Fuente Numerador", "Registro Nacional de Inmunizaciones (RNI) — Bases Programáticas por Residencia"),
        ("Fuente Denominador", "Nacidos Vivos DEIS (por residencia) menos Defunciones acumuladas 0-14 años DEIS (por residencia)"),
        ("Vacunas Consideradas", "Todas las formulaciones VPH (Tetravalente, Nonavalente, Bivalente) del sector público y privado"),
        ("Estratificación", "Por Año de Medición (2019-2026), por Sexo (Mujeres, Hombres, Total) y por Comuna de Residencia"),
        ("Línea Base Nacional (2020)", "74.0%"),
        ("Meta 2030 (Plan Nacional Cáncer)", "90.0%"),
        ("Fecha de Corte Año 2026", "31/07/2026 (Único año en desarrollo / incompleto)"),
    ]

    r = 4
    for label, val in ficha_items:
        ws1.cell(r, 1, label).font = font_bold
        ws1.cell(r, 1).fill = fill_zebra
        ws1.cell(r, 1).border = thin_border
        ws1.cell(r, 1).alignment = align_left
        ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        ws1.cell(r, 2, val).font = font_data
        ws1.cell(r, 2).border = thin_border
        ws1.cell(r, 2).alignment = align_left
        r += 1

    r += 1
    ws1.merge_cells(f"A{r}:G{r}")
    ws1[f"A{r}"] = "EVOLUCIÓN HISTÓRICA DEL ESQUEMA DE VACUNACIÓN VPH EN EL PNI (CHILE)"
    ws1[f"A{r}"].font = font_section
    ws1[f"A{r}"].fill = PatternFill(start_color="ED8936", end_color="ED8936", fill_type="solid")
    ws1[f"A{r}"].alignment = align_center
    r += 1

    headers_evol = ["Período", "Población Objetivo Escolar", "Esquema / Dosis", "Vacuna Utilizada", "Criterio Esquema Completo a los 15 Años"]
    ws1.row_dimensions[r].height = 25
    col_spans = [(1, 1), (2, 3), (4, 4), (5, 5), (6, 7)]
    
    # Headers
    ws1.cell(r, 1, "Período / Año").font = font_tbl_header
    ws1.cell(r, 1).fill = fill_primary
    ws1.merge_cells(f"B{r}:C{r}")
    ws1.cell(r, 2, "Población Objetivo").font = font_tbl_header
    ws1.cell(r, 2).fill = fill_primary
    ws1.cell(r, 4, "Esquema Escolar").font = font_tbl_header
    ws1.cell(r, 4).fill = fill_primary
    ws1.cell(r, 5, "Vacuna PNI").font = font_tbl_header
    ws1.cell(r, 5).fill = fill_primary
    ws1.merge_cells(f"F{r}:G{r}")
    ws1.cell(r, 6, "Criterio Esquema Completo a los 15 Años").font = font_tbl_header
    ws1.cell(r, 6).fill = fill_primary
    r += 1

    filas_evol = [
        ("2014 - 2018", "Solo Mujeres (Niñas de 4° y 5° Básico)", "2 Dosis (4° Básico 1ª dosis + 5° Básico 2ª dosis)", "VPH Tetravalente", "2+ dosis registradas (o dosis única)"),
        ("2019", "Mujeres (4° y 5° Básico) e Incorporación de Hombres (4° Básico)", "Niñas: 1ª y 2ª dosis / Niños: 1ª dosis en 4° Básico", "VPH Tetravalente", "2+ dosis registradas (o dosis única)"),
        ("2020 - 2023", "Ambos Sexos (Niñas y Niños de 4° y 5° Básico)", "2 Dosis para ambos sexos (4° Básico 1ª + 5° Básico 2ª)", "VPH Tetravalente", "2+ dosis registradas (o dosis única)"),
        ("2024 en adelante", "Ambos Sexos (Niñas y Niños de 4° Básico)", "Dosis Única (Nuevo calendario nacional)", "VPH Nonavalente", "1+ dosis registrada (Dosis Única Nonavalente)"),
    ]

    for per, obj, esq, vac, crit in filas_evol:
        ws1.cell(r, 1, per).font = font_bold
        ws1.cell(r, 1).alignment = align_center
        ws1.cell(r, 1).border = thin_border

        ws1.merge_cells(f"B{r}:C{r}")
        ws1.cell(r, 2, obj).font = font_data
        ws1.cell(r, 2).alignment = align_left
        ws1.cell(r, 2).border = thin_border
        ws1.cell(r, 3).border = thin_border

        ws1.cell(r, 4, esq).font = font_data
        ws1.cell(r, 4).alignment = align_left
        ws1.cell(r, 4).border = thin_border

        ws1.cell(r, 5, vac).font = font_data
        ws1.cell(r, 5).alignment = align_center
        ws1.cell(r, 5).border = thin_border

        ws1.merge_cells(f"F{r}:G{r}")
        ws1.cell(r, 6, crit).font = font_bold
        ws1.cell(r, 6).alignment = align_left
        ws1.cell(r, 6).border = thin_border
        ws1.cell(r, 7).border = thin_border
        r += 1

    # Anchos de columna Hoja 1
    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 25
    ws1.column_dimensions["C"].width = 20
    ws1.column_dimensions["D"].width = 30
    ws1.column_dimensions["E"].width = 22
    ws1.column_dimensions["F"].width = 25
    ws1.column_dimensions["G"].width = 20

    # =========================================================================
    # FUNCIÓN AUXILIAR PARA GENERAR HOJAS DE MATRIZ DE INDICADOR (Total, Mujeres, Hombres)
    # =========================================================================
    def generar_hoja_indicador(sheet_title, filtro_sexo, titulo_principal, fill_color):
        ws = wb.create_sheet(title=sheet_title)
        ws.views.sheetView[0].showGridLines = True

        ws.merge_cells("A1:Y1")
        ws["A1"] = titulo_principal
        ws["A1"].font = font_title
        ws["A1"].fill = fill_color
        ws["A1"].alignment = align_center
        ws.row_dimensions[1].height = 30

        ws.merge_cells("A2:Y2")
        ws["A2"] = "Criterio de Residencia: Numerador (RNI) y Denominador (Nacidos Vivos DEIS - Defunciones 0-14 años DEIS). Serie Histórica 2019-2026."
        ws["A2"].font = font_subtitle
        ws["A2"].fill = fill_secondary
        ws["A2"].alignment = align_center
        ws.row_dimensions[2].height = 20

        # Encabezados de tabla
        ws.merge_cells("A4:A5")
        ws["A4"] = "Cód."
        ws["A4"].font = font_tbl_header
        ws["A4"].fill = fill_primary
        ws["A4"].alignment = align_center
        ws["A4"].border = thin_border
        ws["A5"].border = thin_border

        ws.merge_cells("B4:B5")
        ws["B4"] = "Comuna de Residencia"
        ws["B4"].font = font_tbl_header
        ws["B4"].fill = fill_primary
        ws["B4"].alignment = align_center
        ws["B4"].border = thin_border
        ws["B5"].border = thin_border

        col_idx = 3
        for ano in ANOS_REPORTE:
            c = COHORTES_NACIMIENTO[ano]
            col_letter_start = get_column_letter(col_idx)
            col_letter_end = get_column_letter(col_idx + 2)
            ws.merge_cells(f"{col_letter_start}4:{col_letter_end}4")
            
            lbl = f"Año {ano}\n(Cohorte {c})" + (" (*)" if ano == 2026 else "")
            cell = ws[f"{col_letter_start}4"]
            cell.value = lbl
            cell.font = font_tbl_header
            cell.fill = fill_primary
            cell.alignment = align_center
            cell.border = thin_border

            subheaders = ["Num.", "Den.", "Cob. %"]
            for s_idx, sh in enumerate(subheaders):
                c_cell = ws.cell(5, col_idx + s_idx, sh)
                c_cell.font = font_tbl_header
                c_cell.fill = fill_secondary
                c_cell.alignment = align_center
                c_cell.border = thin_border

            col_idx += 3

        ws.row_dimensions[4].height = 26
        ws.row_dimensions[5].height = 20

        # Filas de datos por comuna
        row_cur = 6
        for cod_com in CODIGOS_COMUNAS:
            nom_com = COMUNAS_OSORNO[cod_com]
            ws.cell(row_cur, 1, cod_com).font = font_bold
            ws.cell(row_cur, 1).alignment = align_center
            ws.cell(row_cur, 1).border = thin_border

            ws.cell(row_cur, 2, nom_com).font = font_data
            ws.cell(row_cur, 2).alignment = align_left
            ws.cell(row_cur, 2).border = thin_border

            c_idx = 3
            for ano in ANOS_REPORTE:
                d = resultados[ano][cod_com][filtro_sexo]
                ws.cell(row_cur, c_idx, d["num"]).number_format = "#,##0"
                ws.cell(row_cur, c_idx).alignment = align_right
                ws.cell(row_cur, c_idx).border = thin_border

                ws.cell(row_cur, c_idx + 1, d["den"]).number_format = "#,##0"
                ws.cell(row_cur, c_idx + 1).alignment = align_right
                ws.cell(row_cur, c_idx + 1).border = thin_border

                c_pct = ws.cell(row_cur, c_idx + 2, d["pct"] / 100.0)
                c_pct.number_format = "0.00%"
                c_pct.alignment = align_right
                c_pct.border = thin_border
                c_pct.font = font_bold

                c_idx += 3

            row_cur += 1

        # Fila TOTAL S.S. OSORNO
        ws.cell(row_cur, 1, "23").font = font_bold
        ws.cell(row_cur, 1).alignment = align_center
        ws.cell(row_cur, 1).fill = fill_total
        ws.cell(row_cur, 1).border = total_border

        ws.cell(row_cur, 2, "TOTAL S.S. OSORNO").font = font_bold
        ws.cell(row_cur, 2).alignment = align_left
        ws.cell(row_cur, 2).fill = fill_total
        ws.cell(row_cur, 2).border = total_border

        c_idx = 3
        for ano in ANOS_REPORTE:
            d = resultados[ano]["TOTAL"][filtro_sexo]
            c1 = ws.cell(row_cur, c_idx, d["num"])
            c1.number_format = "#,##0"
            c1.font = font_bold
            c1.alignment = align_right
            c1.fill = fill_total
            c1.border = total_border

            c2 = ws.cell(row_cur, c_idx + 1, d["den"])
            c2.number_format = "#,##0"
            c2.font = font_bold
            c2.alignment = align_right
            c2.fill = fill_total
            c2.border = total_border

            c3 = ws.cell(row_cur, c_idx + 2, d["pct"] / 100.0)
            c3.number_format = "0.00%"
            c3.font = font_bold
            c3.alignment = align_right
            c3.fill = fill_total
            c3.border = total_border

            c_idx += 3

        row_cur += 2
        # Notas al pie
        ws.merge_cells(f"A{row_cur}:Y{row_cur}")
        ws[f"A{row_cur}"] = "(*) NOTA TÉCNICA AÑO 2026: Datos parciales del año en desarrollo (único año incompleto). Fecha de corte de la base RNI: 31/07/2026."
        ws[f"A{row_cur}"].font = font_note
        ws[f"A{row_cur}"].fill = fill_note

        row_cur += 1
        ws.merge_cells(f"A{row_cur}:Y{row_cur}")
        ws[f"A{row_cur}"] = "NOTA EPIDEMIOLÓGICA SEXO MASCULINO: La vacunación escolar de hombres se inició en 2019 con 4° Básico (cohorte 2009). Por ende, las cohortes masculinas 2004-2008 no fueron vacunadas programáticamente en el colegio a los 9-10 años, explicando sus coberturas históricas."
        ws[f"A{row_cur}"].font = font_note

        # Ajuste de anchos
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 24
        for c_i in range(3, col_idx):
            col_letter = get_column_letter(c_i)
            ws.column_dimensions[col_letter].width = 11

    # =========================================================================
    # GENERAR HOJAS 2, 3 Y 4
    # =========================================================================
    generar_hoja_indicador(
        "Indicador 15 Años - Total",
        "Total",
        "COBERTURA DE VACUNACIÓN VPH A LOS 15 AÑOS — TOTAL AMBOS SEXOS (2019-2026)",
        fill_primary
    )

    generar_hoja_indicador(
        "Indicador 15 Años - Mujeres",
        "Mujer",
        "COBERTURA DE VACUNACIÓN VPH A LOS 15 AÑOS — SEXO FEMENINO (MUJERES) (2019-2026)",
        fill_fem
    )

    generar_hoja_indicador(
        "Indicador 15 Años - Hombres",
        "Hombre",
        "COBERTURA DE VACUNACIÓN VPH A LOS 15 AÑOS — SEXO MASCULINO (HOMBRES) (2019-2026)",
        fill_masc
    )

    # =========================================================================
    # HOJA 5: DOSIS ANUALES POR COMUNA Y SEXO
    # =========================================================================
    ws5 = wb.create_sheet(title="Dosis Anuales por Comuna y Sexo")
    ws5.views.sheetView[0].showGridLines = True

    ws5.merge_cells("A1:H1")
    ws5["A1"] = "DOSIS DE VACUNA VPH ADMINISTRADAS POR AÑO, COMUNA DE RESIDENCIA Y SEXO (2015-2026)"
    ws5["A1"].font = font_title
    ws5["A1"].fill = fill_primary
    ws5["A1"].alignment = align_center
    ws5.row_dimensions[1].height = 30

    ws5.merge_cells("A2:H2")
    ws5["A2"] = "Fuente: Bases Programáticas RNI por Residencia. Desglose anual según cambios en el calendario de vacunación."
    ws5["A2"].font = font_subtitle
    ws5["A2"].fill = fill_secondary
    ws5["A2"].alignment = align_center
    ws5.row_dimensions[2].height = 20

    headers_ws5 = ["Año Admin.", "Cód. Comuna", "Comuna Residencia", "Sexo", "1ª Dosis", "2ª Dosis", "Dosis Única", "Total Dosis"]
    for c_i, h_txt in enumerate(headers_ws5, start=1):
        cell = ws5.cell(4, c_i, h_txt)
        cell.font = font_tbl_header
        cell.fill = fill_primary
        cell.alignment = align_center
        cell.border = thin_border
    ws5.row_dimensions[4].height = 22

    row_ws5 = 5
    for ano in sorted(PROGRAMATICAS_FILES.keys()):
        for cod_com in CODIGOS_COMUNAS + ["TOTAL"]:
            nom_com = COMUNAS_OSORNO.get(cod_com, "TOTAL S.S. OSORNO")
            for sexo in ["Mujer", "Hombre", "Total"]:
                d_dict = dosis_anuales[ano][cod_com][sexo]
                d1 = d_dict["1ª Dosis"]
                d2 = d_dict["2ª Dosis"]
                du = d_dict["Dosis Única"]
                dtot = d_dict["Total"]

                if dtot == 0 and cod_com != "TOTAL":
                    continue

                ws5.cell(row_ws5, 1, ano).alignment = align_center
                ws5.cell(row_ws5, 1).border = thin_border

                ws5.cell(row_ws5, 2, cod_com).alignment = align_center
                ws5.cell(row_ws5, 2).border = thin_border

                ws5.cell(row_ws5, 3, nom_com).alignment = align_left
                ws5.cell(row_ws5, 3).border = thin_border

                ws5.cell(row_ws5, 4, sexo).alignment = align_center
                ws5.cell(row_ws5, 4).border = thin_border

                ws5.cell(row_ws5, 5, d1).number_format = "#,##0"
                ws5.cell(row_ws5, 5).alignment = align_right
                ws5.cell(row_ws5, 5).border = thin_border

                ws5.cell(row_ws5, 6, d2).number_format = "#,##0"
                ws5.cell(row_ws5, 6).alignment = align_right
                ws5.cell(row_ws5, 6).border = thin_border

                ws5.cell(row_ws5, 7, du).number_format = "#,##0"
                ws5.cell(row_ws5, 7).alignment = align_right
                ws5.cell(row_ws5, 7).border = thin_border

                c_tot = ws5.cell(row_ws5, 8, dtot)
                c_tot.number_format = "#,##0"
                c_tot.font = font_bold
                c_tot.alignment = align_right
                c_tot.border = thin_border

                if cod_com == "TOTAL" and sexo == "Total":
                    for c_k in range(1, 9):
                        ws5.cell(row_ws5, c_k).fill = fill_total
                        ws5.cell(row_ws5, c_k).border = total_border

                row_ws5 += 1

    ws5.column_dimensions["A"].width = 12
    ws5.column_dimensions["B"].width = 14
    ws5.column_dimensions["C"].width = 24
    ws5.column_dimensions["D"].width = 12
    ws5.column_dimensions["E"].width = 12
    ws5.column_dimensions["F"].width = 12
    ws5.column_dimensions["G"].width = 14
    ws5.column_dimensions["H"].width = 14

    # =========================================================================
    # HOJA 6: DETALLE DENOMINADORES DEIS
    # =========================================================================
    ws6 = wb.create_sheet(title="Detalle Denominadores DEIS")
    ws6.views.sheetView[0].showGridLines = True

    ws6.merge_cells("A1:H1")
    ws6["A1"] = "DESCOMPOSICIÓN DE DENOMINADORES POR COHORTE, COMUNA DE RESIDENCIA Y SEXO"
    ws6["A1"].font = font_title
    ws6["A1"].fill = fill_primary
    ws6["A1"].alignment = align_center
    ws6.row_dimensions[1].height = 30

    ws6.merge_cells("A2:H2")
    ws6["A2"] = "Fórmula: Denominador Neto = Nacidos Vivos DEIS (Residencia) - Defunciones acumuladas 0-14 años DEIS (Residencia)"
    ws6["A2"].font = font_subtitle
    ws6["A2"].fill = fill_secondary
    ws6["A2"].alignment = align_center
    ws6.row_dimensions[2].height = 20

    headers_ws6 = ["Año Medición", "Cohorte (Nac.)", "Cód. Comuna", "Comuna Residencia", "Sexo", "Nacidos Vivos DEIS", "Defunciones 0-14 DEIS", "Denominador Neto"]
    for c_i, h_txt in enumerate(headers_ws6, start=1):
        cell = ws6.cell(4, c_i, h_txt)
        cell.font = font_tbl_header
        cell.fill = fill_primary
        cell.alignment = align_center
        cell.border = thin_border
    ws6.row_dimensions[4].height = 22

    row_ws6 = 5
    for ano in ANOS_REPORTE:
        c = COHORTES_NACIMIENTO[ano]
        for cod_com in CODIGOS_COMUNAS + ["TOTAL"]:
            nom_com = COMUNAS_OSORNO.get(cod_com, "TOTAL S.S. OSORNO")
            for sexo in ["Total", "Mujer", "Hombre"]:
                d = resultados[ano][cod_com][sexo]
                ws6.cell(row_ws6, 1, ano).alignment = align_center
                ws6.cell(row_ws6, 1).border = thin_border

                ws6.cell(row_ws6, 2, c).alignment = align_center
                ws6.cell(row_ws6, 2).border = thin_border

                ws6.cell(row_ws6, 3, cod_com).alignment = align_center
                ws6.cell(row_ws6, 3).border = thin_border

                ws6.cell(row_ws6, 4, nom_com).alignment = align_left
                ws6.cell(row_ws6, 4).border = thin_border

                ws6.cell(row_ws6, 5, sexo).alignment = align_center
                ws6.cell(row_ws6, 5).border = thin_border

                ws6.cell(row_ws6, 6, d["nac"]).number_format = "#,##0"
                ws6.cell(row_ws6, 6).alignment = align_right
                ws6.cell(row_ws6, 6).border = thin_border

                ws6.cell(row_ws6, 7, d["def"]).number_format = "#,##0"
                ws6.cell(row_ws6, 7).alignment = align_right
                ws6.cell(row_ws6, 7).border = thin_border

                c_neto = ws6.cell(row_ws6, 8, d["den"])
                c_neto.number_format = "#,##0"
                c_neto.font = font_bold
                c_neto.alignment = align_right
                c_neto.border = thin_border

                if cod_com == "TOTAL" and sexo == "Total":
                    for c_k in range(1, 9):
                        ws6.cell(row_ws6, c_k).fill = fill_total
                        ws6.cell(row_ws6, c_k).border = total_border

                row_ws6 += 1

    ws6.column_dimensions["A"].width = 13
    ws6.column_dimensions["B"].width = 14
    ws6.column_dimensions["C"].width = 14
    ws6.column_dimensions["D"].width = 24
    ws6.column_dimensions["E"].width = 12
    ws6.column_dimensions["F"].width = 18
    ws6.column_dimensions["G"].width = 20
    ws6.column_dimensions["H"].width = 18

    # =========================================================================
    # HOJA 7: DETALLE ESTABLECIMIENTOS
    # =========================================================================
    ws7 = wb.create_sheet(title="Detalle Establecimientos")
    ws7.views.sheetView[0].showGridLines = True

    headers_estab = ["Año Medición", "Cód. Comuna", "Comuna Residencia", "Establecimiento Vacunación", "Personas Esquema Completo"]
    for c_i, h_txt in enumerate(headers_estab, start=1):
        cell = ws7.cell(1, c_i, h_txt)
        cell.font = font_tbl_header
        cell.fill = fill_primary
        cell.alignment = align_center
        cell.border = thin_border
    ws7.row_dimensions[1].height = 24

    r_estab = 2
    for (ano_m, cod_c, nom_estab), cant in sorted(estab_detalle.items()):
        ws7.cell(r_estab, 1, ano_m).alignment = align_center
        ws7.cell(r_estab, 1).border = thin_border

        ws7.cell(r_estab, 2, cod_c).alignment = align_center
        ws7.cell(r_estab, 2).border = thin_border

        ws7.cell(r_estab, 3, COMUNAS_OSORNO.get(cod_c, cod_c)).alignment = align_left
        ws7.cell(r_estab, 3).border = thin_border

        ws7.cell(r_estab, 4, nom_estab).alignment = align_left
        ws7.cell(r_estab, 4).border = thin_border

        ws7.cell(r_estab, 5, cant).number_format = "#,##0"
        ws7.cell(r_estab, 5).alignment = align_right
        ws7.cell(r_estab, 5).border = thin_border
        r_estab += 1

    ws7.column_dimensions["A"].width = 14
    ws7.column_dimensions["B"].width = 14
    ws7.column_dimensions["C"].width = 24
    ws7.column_dimensions["D"].width = 45
    ws7.column_dimensions["E"].width = 24

    # Guardar archivo
    wb.save(output_path)
    print(f"\n[OK] Reporte Excel guardado exitosamente en:\n  {output_path}")


def main():
    t_inicio = time.time()
    print("=" * 70)
    print("CÁLCULO INTEGRAL DEL INDICADOR VPH A LOS 15 AÑOS — S.S. OSORNO")
    print(f"Inicio: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 70)

    # Paso 1: Leer bases programáticas
    registros_personas, dosis_anuales = leer_programaticas_vph()

    # Paso 2: Determinar esquemas completos
    completos, estab_detalle = determinar_esquema_completo(registros_personas)

    # Paso 3: Leer nacimientos DEIS por residencia y sexo
    nacimientos = leer_nacimientos_por_cohorte()

    # Paso 4: Leer defunciones DEIS 0-14 años por residencia y sexo
    defunciones = leer_defunciones_por_cohorte()

    # Paso 5: Calcular indicadores consolidados
    resultados = calcular_indicadores_completos(completos, nacimientos, defunciones)

    # Paso 6: Generar libro Excel consolidado
    output_excel = os.path.join(BASE_DIR, "Indicador_VPH_15anios_Osorno_2021-2026.xlsx")
    crear_reporte_excel_completo(resultados, dosis_anuales, estab_detalle, output_excel)

    t_total = time.time() - t_inicio
    print(f"\n[T] Tiempo total de ejecución: {t_total:.1f} segundos")
    print("[OK] Proceso finalizado exitosamente.")


if __name__ == "__main__":
    main()
