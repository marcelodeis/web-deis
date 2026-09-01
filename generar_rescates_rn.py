"""
╔══════════════════════════════════════════════════════════════════════════════╗
║   GENERADOR DE RESCATES RN — BCG Y HEPATITIS B (MULTI-AÑO)                 ║
║   Cruce: NAC{Y} × Programáticas_Ocurrencia_{Y} y {Y+1} × DEF               ║
║   Genera Excels protegidos (contraseña: DEIS2026) y JSON web.              ║
╚══════════════════════════════════════════════════════════════════════════════╝
"""

import csv
import os
import json
from collections import defaultdict
from datetime import datetime

# ── CONFIGURACIÓN ──────────────────────────────────────────────────────────

BASE_DIR = r"C:\Antigravity IDE\WEB DEIS\BASE DATOS MINSAL"
OUTPUT_DIR = r"C:\Antigravity IDE\WEB DEIS\Programáticas_Web"
PASSWORD = "DEIS2026"
YEARS = [2020, 2021, 2022, 2023, 2024, 2025, 2026]

# Establecimientos SSO (comparación case-insensitive)
ESTAB_SSO = {
    "CLINICA ALEMANA OSORNO",
    "CLINICA ALEMANA DE OSORNO",
    "HOSPITAL BASE DE OSORNO",
    "HOSPITAL DE OSORNO",
    "CONSULTORIO RURAL HOSP PUERTO OCTAY",
    "HOSPITAL DE PURRANQUE",
    "HOSPITAL RIO NEGRO",
    "CESFAM BAHIA MANSA",
}

# Vacunas a buscar
VACUNAS_BCG = {"BCG_maternidad", "Vacuna BCG"}
VACUNAS_HEPB = {"Hepatitis B pediátrica", "Hepatitis B pediátrica (sector privado)", "Hepatitis B_maternidad"}

# RUNs excluidos manualmente (falsos positivos, vacunados fuera de la red, etc.)
RUNS_EXCLUIDOS = {
    "29192378", # Vacunado BCG en Curicó
}

# Columnas de NAC a incluir en el Excel de salida
NAC_COLS_OUTPUT = [
    "RUN", "NOMBRE", "SEXO", "DIA_NAC", "MES_NAC", "ANO_NAC",
    "PESO", "TALLA", "SEMANAS", "ESTAB", "COMUNA", "DOM_COMUNA",
    "RUN_M", "FOLIO"
]


def clean_run(run_str):
    """Limpia un RUN: quita puntos, guiones, espacios. Devuelve solo dígitos (sin DV)."""
    if not run_str:
        return ""
    s = str(run_str).strip().replace(".", "").replace("-", "").replace(" ", "")
    if not s:
        return ""
    # Si termina en K/k, quitar DV
    if s.upper().endswith("K"):
        return s[:-1]
    # Si es todo dígitos, en NAC ya viene sin DV
    return s


def calcular_dv(rut_num):
    """Calcula el dígito verificador módulo 11 para un RUT numérico."""
    if not rut_num or not str(rut_num).isdigit():
        return ""
    rut = str(rut_num)
    revertido = map(int, reversed(rut))
    factors = [2, 3, 4, 5, 6, 7]
    s = sum(d * factors[i % 6] for i, d in enumerate(revertido))
    dv = 11 - (s % 11)
    if dv == 11:
        return "0"
    elif dv == 10:
        return "K"
    return str(dv)


def formatear_rut(rut_clean):
    """Convierte un RUT limpio (sin DV) al formato XX.XXX.XXX-X."""
    if not rut_clean or not str(rut_clean).isdigit():
        return rut_clean
    dv = calcular_dv(rut_clean)
    rut_fmt = f"{int(rut_clean):,}".replace(",", ".")
    return f"{rut_fmt}-{dv}"


def clean_run_quitar_dv(run_str):
    """Para la base de programáticas donde el RUN incluye DV."""
    if not run_str:
        return ""
    s = str(run_str).strip().replace(".", "").replace("-", "").replace(" ", "")
    if not s:
        return ""
    if s.upper().endswith("K"):
        return s[:-1]
    if s.isdigit() and len(s) > 1:
        # Asumimos último dígito es DV
        return s[:-1]
    return s


def formatear_nombre(nombre_raw):
    """Convierte APELLIDO1/APELLIDO2/NOMBRES= a formato legible."""
    if not nombre_raw:
        return ""
    s = str(nombre_raw).strip().rstrip("=")
    parts = s.split("/")
    if len(parts) >= 3:
        return f"{parts[2].strip()} {parts[0].strip()} {parts[1].strip()}"
    return s


def leer_nacidos_vivos(year):
    """Lee NAC{year}.csv y retorna registros filtrados por ESTAB SSO."""
    nac_file = os.path.join(BASE_DIR, str(year), f"NAC{year}.csv")
    if not os.path.exists(nac_file):
        print(f"  [WARN] No existe archivo NAC para el año {year}: {nac_file}")
        return []
        
    print(f"  Leyendo nacidos vivos: {os.path.basename(nac_file)}")
    nacidos = []
    total = 0
    
    with open(nac_file, encoding="latin1") as f:
        reader = csv.DictReader(f, delimiter=";")
        for row in reader:
            total += 1
            estab = row.get("ESTAB", "").strip()
            if estab.upper() in ESTAB_SSO:
                # Limpiar datos
                registro = {}
                for col in NAC_COLS_OUTPUT:
                    registro[col] = row.get(col, "").strip()
                
                registro["RUN_CLEAN"] = clean_run(registro["RUN"])
                registro["RUN_M_CLEAN"] = clean_run(registro["RUN_M"])
                registro["RUN_FORMATTED"] = formatear_rut(registro["RUN_CLEAN"])
                registro["RUN_M_FORMATTED"] = formatear_rut(registro["RUN_M_CLEAN"])
                registro["NOMBRE_FORMATEADO"] = formatear_nombre(registro["NOMBRE"])
                
                # Construir fecha de nacimiento legible
                dia = registro.get("DIA_NAC", "")
                mes = registro.get("MES_NAC", "")
                ano = registro.get("ANO_NAC", "")
                if dia and mes and ano:
                    registro["FECHA_NACIMIENTO"] = f"{dia.zfill(2)}/{mes.zfill(2)}/{ano}"
                else:
                    registro["FECHA_NACIMIENTO"] = ""
                
                nacidos.append(registro)
    
    print(f"    Total NAC{year}: {total:,}")
    print(f"    Nacidos en establecimientos SSO: {len(nacidos):,}")
    return nacidos


def es_estab_sso_maternidad(estab_nombre):
    """Verifica si el establecimiento en Programáticas corresponde a la red SSO señalada."""
    e = estab_nombre.upper()
    return (
        ("SAN JOS" in e and "OSORNO" in e) or
        ("HOSPITAL" in e and "OSORNO" in e) or
        ("ALEMANA" in e and "OSORNO" in e) or
        ("HOSPITAL" in e and "PURRANQUE" in e) or
        ("HOSPITAL" in e and "NEGRO" in e) or
        ("HOSPITAL" in e and "OCTAY" in e) or
        ("BAH" in e and "MANSA" in e)
    )

def leer_vacunados_programaticas(year):
    """
    Lee Programáticas_Ocurrencia de {year} y {year+1} para capturar vacunas rezagadas.
    """
    archivos_a_leer = [os.path.join(BASE_DIR, str(year), f"Programáticas_Ocurrencia_{year}.csv")]
    
    if year < max(YEARS):
        archivos_a_leer.append(os.path.join(BASE_DIR, str(year+1), f"Programáticas_Ocurrencia_{year+1}.csv"))
        
    vacunados_bcg = set()
    vacunados_hepb = set()
    rechazos_bcg = []
    rechazos_hepb = []
    rechazos_bcg_causas = {} # map RUN -> CAUSA
    rechazos_hepb_causas = {} # map RUN -> CAUSA
    
    def mapear_causa(c):
        c_upper = c.strip().upper()
        if not c_upper or c_upper == "SIN CAUSA REGISTRADA": return "Sin causa registrada"
        if "PADRES" in c_upper or "PACIENTE" in c_upper: return "Solicitud del padre/madre o responsable"
        if "FALLECIMIENTO" in c_upper: return "Fallecimiento"
        if "CONTRAINDICACION" in c_upper or "CONTRA" in c_upper: return "Contraindicación médica"
        return c.strip()
    
    total = 0
    total_bcg = 0
    total_hepb = 0
    
    for archivo in archivos_a_leer:
        if not os.path.exists(archivo):
            print(f"    [WARN] No se encontró el archivo {os.path.basename(archivo)}")
            continue
            
        print(f"  Leyendo vacunas programáticas: {os.path.basename(archivo)}")
        
        with open(archivo, encoding="latin1") as f:
            reader = csv.DictReader(f, delimiter="|")
            for row in reader:
                total += 1
                
                nombre_vacuna = row.get("NOMBRE_VACUNA", "").strip()
                
                # Determinar si es BCG o HepB
                es_bcg = nombre_vacuna in VACUNAS_BCG
                es_hepb = nombre_vacuna in VACUNAS_HEPB
                
                if not es_bcg and not es_hepb:
                    continue
                
                run_raw = row.get("RUN", "").strip()
                run_clean = clean_run_quitar_dv(run_raw)
                
                # Verificar filtros DEIS obligatorios
                vac_admin = row.get("VACUNA_ADMINISTRADA", "").strip().upper()
                reg_elim = row.get("REGISTRO_ELIMINADO", "").strip().upper()
                criterio = row.get("CRITERIO_ELEGIBILIDAD", "").strip().upper()
                dosis = row.get("DOSIS", "").strip().upper()
                
                # Rechazos
                if vac_admin != "SI":
                    if reg_elim == "NO" and criterio != "EPRO" and dosis != "EPRO":
                        estab_ocurrencia = row.get("ESTABLECIMIENTO", "").strip()
                        if es_estab_sso_maternidad(estab_ocurrencia):
                            rechazo_info = {
                                "RUN": formatear_rut(run_clean) if run_clean else run_raw,
                                "NOMBRES": row.get("NOMBRES", "").strip(),
                                "APELLIDO_PATERNO": row.get("APELLIDO_PATERNO", "").strip(),
                                "APELLIDO_MATERNO": row.get("APELLIDO_MATERNO", "").strip(),
                                "NOMBRE_VACUNA": nombre_vacuna,
                                "DOSIS": row.get("DOSIS", "").strip(),
                                "CAUSA_RECHAZO": row.get("CAUSA_RECHAZO", "").strip(),
                                "FECHA_INMUNIZACION": row.get("FECHA_INMUNIZACION", "").strip(),
                                "ESTABLECIMIENTO": estab_ocurrencia,
                                "COMUNA_OCURR": row.get("COMUNA_OCURR", "").strip(),
                            }
                            if es_bcg:
                                rechazos_bcg.append(rechazo_info)
                                if run_clean: rechazos_bcg_causas[run_clean] = mapear_causa(row.get("CAUSA_RECHAZO", ""))
                            if es_hepb:
                                rechazos_hepb.append(rechazo_info)
                                if run_clean: rechazos_hepb_causas[run_clean] = mapear_causa(row.get("CAUSA_RECHAZO", ""))
                    continue
                
                # Filtros DEIS
                if reg_elim != "NO":
                    continue
                if criterio == "EPRO":
                    continue
                if dosis == "EPRO":
                    continue
                
                if not run_clean:
                    continue
                
                if es_bcg:
                    if run_clean not in vacunados_bcg:
                        vacunados_bcg.add(run_clean)
                        total_bcg += 1
                if es_hepb:
                    if run_clean not in vacunados_hepb:
                        vacunados_hepb.add(run_clean)
                        total_hepb += 1
    
    # Añadir excluidos manualmente
    for run_excluido in RUNS_EXCLUIDOS:
        if run_excluido not in vacunados_bcg:
            vacunados_bcg.add(run_excluido)
            total_bcg += 1
        if run_excluido not in vacunados_hepb:
            vacunados_hepb.add(run_excluido)
            total_hepb += 1

    print(f"    Total registros leídos: {total:,}")
    print(f"    RUNs vacunados BCG (cruzado): {len(vacunados_bcg):,}")
    print(f"    RUNs vacunados HepB (cruzado): {len(vacunados_hepb):,}")
    print(f"    Rechazos BCG (cruzado): {len(rechazos_bcg):,}")
    print(f"    Rechazos HepB (cruzado): {len(rechazos_hepb):,}")
    
    return vacunados_bcg, vacunados_hepb, rechazos_bcg, rechazos_hepb, rechazos_bcg_causas, rechazos_hepb_causas


def leer_defunciones():
    """Lee todas las bases de defunciones históricas y retorna set de RUNs fallecidos."""
    print("  Leyendo bases de defunciones históricas...")
    fallecidos = set()
    
    archivos_def = []
    for root, _, files in os.walk(BASE_DIR):
        for f in files:
            if f.upper().startswith("DEF") and f.endswith(".csv"):
                archivos_def.append(os.path.join(root, f))
    
    for archivo in sorted(archivos_def):
        try:
            with open(archivo, encoding="latin1") as f:
                reader = csv.DictReader(f, delimiter="|")
                for row in reader:
                    run = row.get("RUN", "").strip()
                    if run:
                        run_clean = clean_run(run)
                        if run_clean:
                            fallecidos.add(run_clean)
        except Exception as e:
            print(f"    Error leyendo {os.path.basename(archivo)}: {e}")
    
    print(f"    Total RUNs fallecidos históricos: {len(fallecidos):,}")
    return fallecidos


def generar_excel(pendientes, rechazos, output_filename, hoja_pendientes, hoja_rechazos):
    """Genera un Excel con 2 hojas: pendientes y rechazos, con formato profesional."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()

    # ── Estilos comunes ──
    thin_border = Border(
        left=Side(style="thin", color="B4C6E7"),
        right=Side(style="thin", color="B4C6E7"),
        top=Side(style="thin", color="B4C6E7"),
        bottom=Side(style="thin", color="B4C6E7")
    )
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    wrap_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── Hoja 1: Pendientes ──
    ws1 = wb.active
    ws1.title = hoja_pendientes

    headers_pendientes = [
        "RUN Menor", "Nombre Completo", "Sexo", "Fecha Nacimiento",
        "Peso (g)", "Talla (cm)", "Semanas Gestación", "Establecimiento de Nacimiento",
        "Cód. Comuna Residencia", "Comuna Residencia", "RUN Madre", "Folio",
        "Estado Búsqueda Madre"
    ]

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    row_fill_even = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    row_fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    data_font = Font(name="Calibri", size=10)

    for col_idx, header in enumerate(headers_pendientes, 1):
        cell = ws1.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap_center
        cell.border = thin_border

    for row_idx, reg in enumerate(pendientes, 2):
        sexo_text = "Masculino" if reg.get("SEXO") == "1" else ("Femenino" if reg.get("SEXO") == "2" else reg.get("SEXO", ""))
        is_even = (row_idx % 2 == 0)
        fill = row_fill_even if is_even else row_fill_odd

        valores = [
            reg.get("RUN_FORMATTED", ""),
            reg.get("NOMBRE_FORMATEADO", ""),
            sexo_text,
            reg.get("FECHA_NACIMIENTO", ""),
            reg.get("PESO", ""),
            reg.get("TALLA", ""),
            reg.get("SEMANAS", ""),
            reg.get("ESTAB", ""),
            reg.get("COMUNA", ""),
            reg.get("DOM_COMUNA", ""),
            reg.get("RUN_M_FORMATTED", ""),
            reg.get("FOLIO", ""),
            reg.get("ESTADO_BUSQUEDA_MADRE", ""),
        ]

        for col_idx, val in enumerate(valores, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.font = data_font
            cell.fill = fill
            if col_idx in [1, 3, 4, 5, 6, 7, 11, 12]:
                cell.alignment = center_align
            else:
                cell.alignment = left_align

    # Auto-ajustar anchos
    for col_idx in range(1, len(headers_pendientes) + 1):
        max_len = len(str(headers_pendientes[col_idx - 1]))
        for r in range(2, min(len(pendientes) + 2, 100)):
            cell_val = ws1.cell(row=r, column=col_idx).value
            if cell_val:
                max_len = max(max_len, len(str(cell_val)))
        ws1.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 3, 40)

    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = ws1.dimensions

    # ── Hoja 2: Rechazos ──
    ws2 = wb.create_sheet(title=hoja_rechazos)

    headers_rechazos = [
        "RUN", "Nombres", "Apellido Paterno", "Apellido Materno",
        "Vacuna", "Dosis", "Causa Rechazo", "Fecha Inmunización",
        "Establecimiento", "Comuna Ocurrencia"
    ]

    rechazo_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
    rechazo_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    rechazo_row_even = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    rechazo_row_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    for col_idx, header in enumerate(headers_rechazos, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.fill = rechazo_fill
        cell.font = rechazo_font
        cell.alignment = wrap_center
        cell.border = thin_border

    for row_idx, rec in enumerate(rechazos, 2):
        is_even = (row_idx % 2 == 0)
        fill = rechazo_row_even if is_even else rechazo_row_odd

        valores = [
            rec.get("RUN", ""),
            rec.get("NOMBRES", ""),
            rec.get("APELLIDO_PATERNO", ""),
            rec.get("APELLIDO_MATERNO", ""),
            rec.get("NOMBRE_VACUNA", ""),
            rec.get("DOSIS", ""),
            rec.get("CAUSA_RECHAZO", ""),
            rec.get("FECHA_INMUNIZACION", ""),
            rec.get("ESTABLECIMIENTO", ""),
            rec.get("COMUNA_OCURR", ""),
        ]
        for col_idx, val in enumerate(valores, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.font = data_font
            cell.fill = fill
            if col_idx in [1, 6]:
                cell.alignment = center_align
            else:
                cell.alignment = left_align

    # Auto-ajustar anchos rechazos
    for col_idx in range(1, len(headers_rechazos) + 1):
        max_len = len(str(headers_rechazos[col_idx - 1]))
        for r in range(2, min(len(rechazos) + 2, 100)):
            cell_val = ws2.cell(row=r, column=col_idx).value
            if cell_val:
                max_len = max(max_len, len(str(cell_val)))
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 3, 40)

    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = ws2.dimensions

    # Guardar
    output_path = os.path.join(OUTPUT_DIR, output_filename)
    wb.save(output_path)
    return output_path


def cifrar_excel(filepath, password):
    """Cifra un archivo Excel con contraseña usando win32com (Excel COM)."""
    try:
        import win32com.client
        excel = win32com.client.Dispatch("Excel.Application")
        excel.DisplayAlerts = False
        excel.Visible = False
        
        abs_path = os.path.abspath(filepath)
        wb = excel.Workbooks.Open(abs_path)
        wb.Password = password
        wb.SaveAs(abs_path, Password=password)
        wb.Close()
        excel.Quit()
        return True
    except Exception as e:
        print(f"    [WARN] No se pudo cifrar {os.path.basename(filepath)}: {e}")
        return False


def main():
    print("=" * 70)
    print("  GENERADOR DE RESCATES RN - BCG Y HEPATITIS B (MULTI-AÑO)")
    print(f"  Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    print("=" * 70)
    
    # Cargar defunciones históricas una sola vez
    fallecidos = leer_defunciones()
    
    data_resumen = {}
    
    for year in sorted(YEARS):
        print(f"\n[{year}] PROCESANDO COHORTE {year}...")
        
        nacidos = leer_nacidos_vivos(year)
        if not nacidos:
            print(f"[{year}] Saltando año. No se encontraron nacidos.")
            continue
            
        vacunados_bcg, vacunados_hepb, rechazos_bcg, rechazos_hepb, rechazos_bcg_causas, rechazos_hepb_causas = leer_vacunados_programaticas(year)
        
        pendientes_bcg = []
        pendientes_hepb = []
        
        stats = {
            "total_nacidos": len(nacidos),
            "bcg_vacunado_run_menor": 0,
            "bcg_vacunado_run_madre": 0,
            "bcg_fallecido": 0,
            "bcg_rechazo": 0,
            "bcg_pendiente": 0,
            "bcg_hospital": defaultdict(int),
            "bcg_hospital_elegibles": defaultdict(int),
            "bcg_comuna": defaultdict(int),
            "bcg_comuna_elegibles": defaultdict(int),
            "bcg_mes": defaultdict(int),
            "bcg_mes_elegibles": defaultdict(int),
            "bcg_causas": defaultdict(int),
            "hepb_vacunado_run_menor": 0,
            "hepb_vacunado_run_madre": 0,
            "hepb_fallecido": 0,
            "hepb_rechazo": 0,
            "hepb_pendiente": 0,
            "hepb_hospital": defaultdict(int),
            "hepb_hospital_elegibles": defaultdict(int),
            "hepb_comuna": defaultdict(int),
            "hepb_comuna_elegibles": defaultdict(int),
            "hepb_mes": defaultdict(int),
            "hepb_mes_elegibles": defaultdict(int),
            "hepb_causas": defaultdict(int)
        }
        
        for reg in nacidos:
            run_menor = reg["RUN_CLEAN"]
            run_madre = reg["RUN_M_CLEAN"]
            estab = reg.get("ESTAB", "Desconocido").strip()
            if not estab: estab = "Desconocido"
            comuna = reg.get("DOM_COMUNA", "Desconocida").strip()
            if not comuna: comuna = "Desconocida"
            mes = reg.get("MES_NAC", "00").strip().zfill(2)
            if not mes or mes == "00": mes = "Desconocido"
            
            if run_menor and run_menor in fallecidos:
                stats["bcg_fallecido"] += 1
                stats["bcg_causas"]["Fallecimiento"] += 1
            elif run_menor and run_menor in rechazos_bcg_causas:
                stats["bcg_rechazo"] += 1
                stats["bcg_causas"][rechazos_bcg_causas[run_menor]] += 1
            else:
                stats["bcg_hospital_elegibles"][estab] += 1
                stats["bcg_comuna_elegibles"][comuna] += 1
                stats["bcg_mes_elegibles"][mes] += 1
                
                vacunado_bcg = False
                if run_menor and run_menor in vacunados_bcg:
                    vacunado_bcg = True
                    stats["bcg_vacunado_run_menor"] += 1
                elif run_madre and run_madre in vacunados_bcg:
                    vacunado_bcg = True
                    stats["bcg_vacunado_run_madre"] += 1
                
                if not vacunado_bcg:
                    reg_copy = dict(reg)
                    reg_copy["ESTADO_BUSQUEDA_MADRE"] = "No encontrado en RUN menor ni madre" if run_madre else "Sin RUN madre"
                    pendientes_bcg.append(reg_copy)
                    stats["bcg_pendiente"] += 1
                    stats["bcg_hospital"][estab] += 1
                    stats["bcg_comuna"][comuna] += 1
                    stats["bcg_mes"][mes] += 1
            
            # HepB
            if run_menor and run_menor in fallecidos:
                stats["hepb_fallecido"] += 1
                stats["hepb_causas"]["Fallecimiento"] += 1
            elif run_menor and run_menor in rechazos_hepb_causas:
                stats["hepb_rechazo"] += 1
                stats["hepb_causas"][rechazos_hepb_causas[run_menor]] += 1
            else:
                stats["hepb_hospital_elegibles"][estab] += 1
                stats["hepb_comuna_elegibles"][comuna] += 1
                stats["hepb_mes_elegibles"][mes] += 1
                
                vacunado_hepb = False
                if run_menor and run_menor in vacunados_hepb:
                    vacunado_hepb = True
                    stats["hepb_vacunado_run_menor"] += 1
                elif run_madre and run_madre in vacunados_hepb:
                    vacunado_hepb = True
                    stats["hepb_vacunado_run_madre"] += 1
                
                if not vacunado_hepb:
                    reg_copy = dict(reg)
                    reg_copy["ESTADO_BUSQUEDA_MADRE"] = "No encontrado en RUN menor ni madre" if run_madre else "Sin RUN madre"
                    pendientes_hepb.append(reg_copy)
                    stats["hepb_pendiente"] += 1
                    stats["hepb_hospital"][estab] += 1
                    stats["hepb_comuna"][comuna] += 1
                    stats["hepb_mes"][mes] += 1
                    
        # El conteo de causas de exclusión exactas se hace arriba en rechazos_bcg_causas y rechazos_hepb_causas

        print(f"  [{year}] BCG: {stats['bcg_vacunado_run_menor']} vacunados / {stats['bcg_pendiente']} pendientes")
        print(f"  [{year}] HepB: {stats['hepb_vacunado_run_menor']} vacunados / {stats['hepb_pendiente']} pendientes")
        
        # Generar Excels
        path_bcg = generar_excel(
            pendientes_bcg, rechazos_bcg,
            output_filename=f"Rescates_BCG_Pendientes_{year}.xlsx",
            hoja_pendientes="Nacidos Vivos sin BCG",
            hoja_rechazos="Rechazos BCG"
        )
        
        path_hepb = generar_excel(
            pendientes_hepb, rechazos_hepb,
            output_filename=f"Rescates_HepB_Pendientes_{year}.xlsx",
            hoja_pendientes="Nacidos Vivos sin HepB",
            hoja_rechazos="Rechazos HepB"
        )
        
        cifrar_excel(path_bcg, PASSWORD)
        cifrar_excel(path_hepb, PASSWORD)
        
        # Guardar resultados para el JSON
        data_resumen[str(year)] = {
            "total_nacidos": stats["total_nacidos"],
            "bcg": {
                "universo": stats["total_nacidos"],
                "excluidos": stats["bcg_fallecido"] + stats["bcg_rechazo"],
                "elegibles": stats["total_nacidos"] - (stats["bcg_fallecido"] + stats["bcg_rechazo"]),
                "vacunados": stats["bcg_vacunado_run_menor"] + stats["bcg_vacunado_run_madre"],
                "pendientes": stats["bcg_pendiente"],
                "fallecidos": stats["bcg_fallecido"],
                "rechazados": stats["bcg_rechazo"],
                "distribucion_hospital": dict(stats["bcg_hospital"]),
                "hospital_elegibles": dict(stats["bcg_hospital_elegibles"]),
                "distribucion_comuna": dict(stats["bcg_comuna"]),
                "evolucion_mensual": dict(stats["bcg_mes"]),
                "mes_elegibles": dict(stats["bcg_mes_elegibles"]),
                "causales_exclusion": dict(stats["bcg_causas"]),
                "calidad_dato": {
                    "rechazos": stats["bcg_rechazo"]
                }
            },
            "hepb": {
                "universo": stats["total_nacidos"],
                "excluidos": stats["hepb_fallecido"] + stats["hepb_rechazo"],
                "elegibles": stats["total_nacidos"] - (stats["hepb_fallecido"] + stats["hepb_rechazo"]),
                "vacunados": stats["hepb_vacunado_run_menor"] + stats["hepb_vacunado_run_madre"],
                "pendientes": stats["hepb_pendiente"],
                "fallecidos": stats["hepb_fallecido"],
                "rechazados": stats["hepb_rechazo"],
                "distribucion_hospital": dict(stats["hepb_hospital"]),
                "hospital_elegibles": dict(stats["hepb_hospital_elegibles"]),
                "distribucion_comuna": dict(stats["hepb_comuna"]),
                "evolucion_mensual": dict(stats["hepb_mes"]),
                "mes_elegibles": dict(stats["hepb_mes_elegibles"]),
                "causales_exclusion": dict(stats["hepb_causas"]),
                "calidad_dato": {
                    "rechazos": stats["hepb_rechazo"]
                }
            }
        }
        
    # Guardar JSON global
    out_json = os.path.join(OUTPUT_DIR, "data_neonatal.json")
    with open(out_json, "w", encoding="utf-8") as f:
        json.dump(data_resumen, f, indent=2, ensure_ascii=False)
        
    out_js = os.path.join(OUTPUT_DIR, "data_neonatal.js")
    with open(out_js, "w", encoding="utf-8") as f:
        f.write("window.dataNeonatal = " + json.dumps(data_resumen, indent=2, ensure_ascii=False) + ";")
        
    print(f"\nGuardado JSON de cobertura histórica: {out_json}")
    print(f"Guardado JS de cobertura histórica: {out_js}")

    print("\n" + "=" * 70)
    print("  PROCESO COMPLETADO")
    print("=" * 70)


if __name__ == "__main__":
    main()
